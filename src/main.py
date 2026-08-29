"""DBI Translation Pipeline.

Usage:
    python -m src.main sync        — sync ua.csv into dictionary.xlsx
    python -m src.main translate   — translate missing cells via AI
    python -m src.main shadok      — localize Shadok parody block via AI
    python -m src.main validate    — validate all translations
    python -m src.main export      — export per-language CSVs
    python -m src.main build       — build .bin files from CSVs
    python -m src.main all         — run full pipeline
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import shutil
import sys
import time
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

# Force stdout to UTF-8 on Windows
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

from src.core.text_utils import tokenize, detokenize, normalize_tokens_out, visual_length, normalize_fullwidth
from src.core.validator import validate
from src.core import ai_client
from src.core.ai_client import translate_batch, refine, init_session, init_session_shadok, translate_shadok_block

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
DICT_PATH = DATA_DIR / "dictionary.xlsx"
UA_CSV = DATA_DIR / "ua.csv"
LANG_JSON = DATA_DIR / "languages.json"
OUTPUT_DIR = ROOT / "output"
TRANSLATIONS_DIR = ROOT / "translations"
DIST_DIR = ROOT / "dist"
BUILD_SCRIPT = ROOT / "scripts" / "build_translation_bin.py"

SHEET_NAME = "Translations"
META_SHEET = "Metadata"

BLOCK_JSON = DATA_DIR / "blocks.json"
SHADOK_JSON = DATA_DIR / "shadok.json"

def find_dbi_version_row(ws, col_map) -> int:
    """Find the row that likely contains the pure 3-4 digit version number."""
    import re as _re
    for row in range(2, min(50, ws.max_row + 1)):
        val = str(ws.cell(row, col_map["Original"]).value or "").strip()
        if _re.match(r"^\d{3,4}$", val):
            return row
    return 9  # fallback


def load_shadok_config() -> dict | None:
    """Load shadok.json config. Returns None if not found."""
    if not SHADOK_JSON.exists():
        return None
    with SHADOK_JSON.open("r", encoding="utf-8") as f:
        return json.load(f)


def get_shadok_target_langs() -> list[str]:
    """Return all languages.json codes except ru."""
    return [lc for lc in load_languages() if lc != "ru"]


def resolve_shadok_mapping_rows(
    ws,
    col_map: dict,
    mapping: list[dict],
) -> list[tuple[int, str, str]]:
    """Resolve mapping items to workbook rows by exact Original==orig (strip-consistent).

    Returns ordered (row_idx, orig, new). Raises ValueError on missing/duplicate orig.
    Does not match by new and does not rewrite Original.
    """
    if "Original" not in col_map:
        raise ValueError("Workbook missing Original column")

    by_orig: dict[str, list[int]] = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_map["Original"]).value
        if val is None:
            continue
        key = str(val).strip()
        if not key:
            continue
        by_orig.setdefault(key, []).append(row)

    resolved: list[tuple[int, str, str]] = []
    for i, item in enumerate(mapping):
        orig = item["orig"]
        new = item["new"]
        key = str(orig).strip()
        rows = by_orig.get(key, [])
        if not rows:
            raise ValueError(
                f"Shadok mapping[{i}] orig not found in workbook: {orig[:80]!r}"
            )
        if len(rows) > 1:
            raise ValueError(
                f"Shadok mapping[{i}] orig matches multiple rows {rows}: {orig[:80]!r}"
            )
        resolved.append((rows[0], orig, new))
    return resolved


def parse_and_validate_shadok_block(
    translated_text: str,
    expected_count: int,
    max_line_length: int,
) -> list[str]:
    """Split/validate a Shadok AI block. Raises ValueError on any contract breach."""
    if translated_text is None:
        raise ValueError("translated_text is None")
    lines = [ln.rstrip("\r") for ln in str(translated_text).split("\n")]
    if len(lines) != expected_count:
        raise ValueError(f"Expected {expected_count} lines, got {len(lines)}")
    for i, line in enumerate(lines):
        if not line.strip():
            raise ValueError(f"Line {i + 1} is empty/whitespace-only")
        vl = visual_length(line)
        if vl > max_line_length:
            raise ValueError(
                f"Line {i + 1} visual_length {vl} > {max_line_length}: {line[:60]!r}"
            )
    return lines


def build_shadok_exclusion_rows(ws, col_map: dict, mapping: list[dict]) -> set[int]:
    """Rows to exclude from general translate/align/structural validate.

    Prefers strict unique resolve. On missing/duplicate mapping, falls back to
    skipping every Original that matches any mapping orig (never matches by new).
    """
    try:
        resolved = resolve_shadok_mapping_rows(ws, col_map, mapping)
        return {row_idx for row_idx, _, _ in resolved}
    except ValueError:
        orig_keys = {str(item["orig"]).strip() for item in mapping}
        rows: set[int] = set()
        if "Original" not in col_map:
            return rows
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, col_map["Original"]).value
            if val is None:
                continue
            if str(val).strip() in orig_keys:
                rows.add(row)
        return rows


def get_nro_version() -> str | None:
    """Extract DBI version from patched NRO filename (e.g. DBI.892.ru_patched.nro -> '892').
    Returns the highest version number found from *_patched.nro files."""
    import re as _re
    versions = []
    for nro_file in ROOT.glob("DBI.*.nro"):
        if "debug" in nro_file.name.lower():
            continue
        match = _re.search(r'DBI\.(\d+)\.', nro_file.name)
        if match and nro_file.name.endswith("_patched.nro"):
            versions.append(int(match.group(1)))

    if versions:
        return str(max(versions))
    return None


def get_patched_nro_path() -> Path | None:
    """Get the path to the latest patched NRO file."""
    import re as _re
    nro_files = [f for f in ROOT.glob("DBI.*.nro") if f.name.endswith("_patched.nro") and "debug" not in f.name.lower()]
    if not nro_files:
        return None

    def extract_version(nro_path):
        match = _re.search(r'DBI\.(\d+)\.', nro_path.name)
        return int(match.group(1)) if match else 0

    return max(nro_files, key=extract_version)

# ── helpers ──────────────────────────────────────────────────────────

def load_languages() -> dict[str, str]:
    with LANG_JSON.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_blocks() -> dict[str, list[str]]:
    if not BLOCK_JSON.exists():
        return {}
    with BLOCK_JSON.open("r", encoding="utf-8") as f:
        return json.load(f)


def open_or_create_workbook() -> openpyxl.Workbook:
    if DICT_PATH.exists():
        return openpyxl.load_workbook(DICT_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    meta = wb.create_sheet(META_SHEET)
    meta["A1"] = "version"
    meta["B1"] = "0.0.0"
    meta["A2"] = "updated"
    meta["B2"] = ""
    return wb


def get_version(wb: openpyxl.Workbook) -> str:
    meta = wb[META_SHEET]
    return str(meta["B1"].value or "0.0.0")


def bump_version(wb: openpyxl.Workbook) -> str:
    meta = wb[META_SHEET]
    ver = str(meta["B1"].value or "0.0.0")
    parts = ver.split(".")
    parts[-1] = str(int(parts[-1]) + 1)
    new_ver = ".".join(parts)
    meta["B1"] = new_ver
    from datetime import datetime, timezone
    meta["B2"] = datetime.now(timezone.utc).isoformat()
    return new_ver


def save_workbook(wb: openpyxl.Workbook) -> None:
    DICT_PATH.parent.mkdir(parents=True, exist_ok=True)
    max_retries = 10
    for attempt in range(max_retries):
        try:
            wb.save(DICT_PATH)
            return
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"\n  [!] File '{DICT_PATH.name}' is locked (likely open in Excel).")
                print(f"      PLEASE CLOSE IT! Retrying in 3s... ({attempt + 1}/{max_retries})")
                time.sleep(3)
            else:
                print(f"\n  [CRITICAL] Could not save '{DICT_PATH.name}' after {max_retries} attempts.")
                raise


def sanitize_string(s: str | None) -> str:
    """Remove illegal control characters that openpyxl cannot handle."""
    if s is None:
        return ""
    s = str(s)
    # Remove control characters except \n, \r, \t
    return re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', s)

# ── translate ────────────────────────────────────────────────────────

def wrap_text(text: str, max_chars: int, lang_code: str) -> list[str]:
    """Wrap text into lines, with special handling for CJK width."""
    # Heuristic for wider CJK characters as requested by user
    effective_max = max_chars
    # Japanese (jp), Korean (kr), Chinese (zh/zhcn/zhtw)
    if lang_code.lower() in ["zhcn", "zhtw", "jp", "kr", "zh"]:
        # "1.5 times smaller"
        effective_max = int(max_chars / 1.5)

    import textwrap
    # replace_whitespace=True converts all tabs/newlines into spaces before wrapping
    lines = textwrap.wrap(text, width=effective_max, break_long_words=True, replace_whitespace=True)
    return lines


def get_translate_workers() -> int:
    """Parse and validate DBI_TRANSLATE_WORKERS environment variable."""
    raw_val = os.environ.get("DBI_TRANSLATE_WORKERS", "4").strip()
    try:
        workers = int(raw_val)
    except ValueError:
        raise ValueError(
            f"Invalid DBI_TRANSLATE_WORKERS value: {raw_val!r}. Expected an integer between 1 and 8."
        )
    if not (1 <= workers <= 8):
        raise ValueError(
            f"Invalid DBI_TRANSLATE_WORKERS value: {workers}. Must be between 1 and 8 (inclusive)."
        )
    return workers


@dataclass
class RowTranslationResult:
    row_id: int
    original: str
    missing: list[str]
    accepted_translations: dict[str, str] = field(default_factory=dict)
    failed_langs: list[str] = field(default_factory=list)
    cyrillic_count: int = 0
    is_skipped_cyrillic: bool = False
    refine_attempts: int = 0
    error_message: Optional[str] = None


def _process_row_translation(row_id: int, original: str, missing: list[str]) -> RowTranslationResult:
    """Pure worker function: performs AI translation, validation, and refine loops for a single row.

    Does NOT touch worksheet or save workbook.
    """
    result = RowTranslationResult(row_id=row_id, original=original, missing=list(missing))

    # Check Cyrillic characters count
    cyrillic_count = len(re.findall(r'[а-яА-ЯёЁіІїЇєЄґҐ]', original))
    result.cyrillic_count = cyrillic_count

    if cyrillic_count < 2:
        result.is_skipped_cyrillic = True
        for lc in missing:
            translation = original
            if re.search(r'\bru\b', original, re.IGNORECASE):
                translation = re.sub(r'\bru\b', lc, original, flags=re.IGNORECASE)
            result.accepted_translations[lc] = translation
        return result

    # Call AI batch translation
    try:
        translations = translate_batch(original, missing, row_id=row_id)
    except Exception as e:
        result.error_message = str(e)
        result.failed_langs = list(missing)
        return result

    # Normalize tokens and full-width chars
    for lc in list(translations.keys()):
        translations[lc] = normalize_fullwidth(normalize_tokens_out(translations.get(lc, "")))

    # Validate + Refine loop
    MAX_REFINE_ATTEMPTS = 3
    for attempt in range(MAX_REFINE_ATTEMPTS):
        errors = []
        for lc in missing:
            translation = translations.get(lc, "")
            if not translation:
                errors.append((lc, "Translation is empty"))
                continue
            ok, msg = validate(original, translation, lc)
            if not ok:
                errors.append((lc, msg))

        if not errors:
            break

        result.refine_attempts = attempt + 1
        if attempt < MAX_REFINE_ATTEMPTS - 1:
            error_lines = "\n".join(f"- {lc}: {msg}" for lc, msg in errors)
            correction = (
                f"The following translations have errors:\n"
                f"{error_lines}\n\n"
                f"Please fix them. Source text: \"{original}\""
            )
            try:
                refined = refine(
                    correction=correction,
                    target_langs=missing,
                    row_id=row_id,
                    original=original,
                    current_translations=translations,
                    validation_errors=errors,
                )
                for lc in list(refined.keys()):
                    refined[lc] = normalize_fullwidth(normalize_tokens_out(refined.get(lc, "")))
                translations.update(refined)
            except Exception:
                break

    # Categorize accepted vs failed
    failed_langs = []
    for lc in missing:
        translation = translations.get(lc, "")
        if not translation or not translation.strip():
            failed_langs.append(lc)
            continue

        if re.search(r'\bru\b', original, re.IGNORECASE):
            translation = re.sub(r'\bru\b', lc, translation, flags=re.IGNORECASE)

        ok, msg = validate(original, translation, lc)
        if ok:
            result.accepted_translations[lc] = translation
        elif "English preservation" in msg:
            result.accepted_translations[lc] = original
        else:
            failed_langs.append(lc)

    # Retry failed languages with error context
    MAX_RETRY_ROUNDS = 3
    for retry_round in range(MAX_RETRY_ROUNDS):
        if not failed_langs:
            break

        error_details = []
        for lc in failed_langs:
            translation = translations.get(lc, "")
            _, msg = validate(original, translation, lc) if translation else (False, "empty")
            error_details.append(f"- {lc}: {msg}")

        error_context = (
            f"Retry round {retry_round + 1}. The following translations for "
            f"\"{original}\" have validation errors:\n"
            + "\n".join(error_details)
            + "\n\nPlease fix these issues. "
            f"Source text: \"{original}\""
        )

        try:
            retry_results = refine(
                correction=error_context,
                target_langs=failed_langs,
                row_id=row_id,
                original=original,
                current_translations={lc: translations.get(lc, "") for lc in failed_langs},
                validation_errors=error_details,
            )
            still_failed = []
            for lc in failed_langs:
                translation = normalize_fullwidth(normalize_tokens_out(retry_results.get(lc, "")))
                if not translation or not translation.strip():
                    still_failed.append(lc)
                    continue

                if re.search(r'\bru\b', original, re.IGNORECASE):
                    translation = re.sub(r'\bru\b', lc, translation, flags=re.IGNORECASE)

                ok, msg = validate(original, translation, lc)
                if ok:
                    result.accepted_translations[lc] = translation
                else:
                    translations[lc] = translation
                    still_failed.append(lc)
            failed_langs = still_failed
        except Exception:
            break

    result.failed_langs = failed_langs
    return result


def cmd_translate() -> None:
    """Find empty cells and translate via AI with continuous chat."""
    # Determine workers limit before starting tasks
    requested_workers = get_translate_workers()
    if ai_client.PROVIDER == "WEB2API" and requested_workers > 1:
        effective_workers = requested_workers
    else:
        effective_workers = 1

    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    force_all = "--force" in sys.argv or "-f" in sys.argv

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    lang_codes = [lc for lc in langs if lc in col_map and lc != "ru"]

    # ── Phase -1: CLEANUP DUPLICATE ROWS ───────────────────────────────
    best_rows = {}
    rows_to_delete = []
    
    for row in range(2, ws.max_row + 1):
        original_val = str(ws.cell(row, col_map["Original"]).value or "")
        key = original_val
        if not key.strip():
            continue
            
        non_empty = 0
        for lc in lang_codes:
            if str(ws.cell(row, col_map[lc]).value or "").strip():
                non_empty += 1
                
        if key in best_rows:
            best_idx, best_count = best_rows[key]
            if non_empty > best_count:
                rows_to_delete.append(best_idx)
                best_rows[key] = (row, non_empty)
            else:
                rows_to_delete.append(row)
        else:
            best_rows[key] = (row, non_empty)
            
    if rows_to_delete:
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        print(f"  [CLEANUP] Deleted {len(rows_to_delete)} exact duplicate rows.")
        save_workbook(wb)

    # ── Phase 0: exclude Shadok rows from general translate ───────────
    shadok_row_set: set[int] = set()
    shadok_config = load_shadok_config()
    if shadok_config:
        shadok_row_set = build_shadok_exclusion_rows(
            ws, col_map, shadok_config.get("mapping", [])
        )

    # ── Scan: count rows that need translation (excluding shadoks) ───
    rows_to_translate = []

    for row in range(2, ws.max_row + 1):
        if row in shadok_row_set:
            continue
            
        original = ws.cell(row, col_map["Original"]).value
        if not original or not str(original).strip():
            continue

        missing = []
        for lc in lang_codes:
            cell_val = ws.cell(row, col_map[lc]).value
            if not cell_val or not str(cell_val).strip():
                missing.append(lc)
            else:
                ok, msg = validate(str(original), str(cell_val), lc)
                if not ok and "English preservation" not in msg:
                    print(f"  [Row {row}][{lc}] Invalid existing translation, scheduling re-translation: {msg}")
                    missing.append(lc)
        if missing:
            rows_to_translate.append((row, original, missing))

    total_rows = len(rows_to_translate)
    total_all = ws.max_row - 1

    if total_rows == 0:
        print(f"  Nothing to translate. All {total_all} rows are complete.")
        return

    # ── Phase 1: INIT ────────────────────────────────────────────────
    print()
    print("=" * 60)
    print("  DBI TRANSLATOR")
    print("=" * 60)
    print(f"  Provider  : {ai_client.PROVIDER} (model: {ai_client.MODEL})")
    print(f"  Workers   : {effective_workers} (configured: {requested_workers})")
    print(f"  Languages : {len(lang_codes)} ({', '.join(lang_codes)})")
    print(f"  Total rows: {total_all}")
    print(f"  To translate: {total_rows} (excl. {len(shadok_row_set)} shadok rows)")
    print("-" * 60)

    init_session()

    print("-" * 60)
    print(f"  Starting translation...")
    print("=" * 60)
    print()

    # ── Phase 2: TRANSLATE ───────────────────────────────────────────
    total_translated = 0
    total_failed = 0

    def apply_row_result(idx: int, res: RowTranslationResult) -> None:
        nonlocal total_translated, total_failed
        row = res.row_id

        if res.is_skipped_cyrillic:
            for lc, translation in res.accepted_translations.items():
                ws.cell(row, col_map[lc], translation)
                total_translated += 1
            save_workbook(wb)
            print(f"    [Row {row} | {idx}/{total_rows}] Cyrillic count {res.cyrillic_count} < 2 — copied as-is.")
            return

        if res.error_message and not res.accepted_translations:
            print(f"  [Row {row} | {idx}/{total_rows}] ERROR: {res.error_message}")
            total_failed += len(res.missing)
            return

        for lc, translation in res.accepted_translations.items():
            ws.cell(row, col_map[lc], translation)
            total_translated += 1

        for lc in res.failed_langs:
            print(f"    [Row {row}][{lc}] SKIPPED (validation failed)")
            total_failed += 1

        save_workbook(wb)

        row_ok = len(res.accepted_translations)
        row_fail = len(res.failed_langs)
        status = "OK" if row_fail == 0 else f"OK:{row_ok} FAIL:{row_fail}"
        print(f"    [Row {row} | {idx}/{total_rows}] Saved. {status}")

    if effective_workers > 1:
        with ThreadPoolExecutor(max_workers=effective_workers) as executor:
            future_to_idx = {
                executor.submit(_process_row_translation, row, original, missing): (idx, row, original, missing)
                for idx, (row, original, missing) in enumerate(rows_to_translate, 1)
            }
            for future in as_completed(future_to_idx):
                idx, row, orig, miss = future_to_idx[future]
                try:
                    res = future.result()
                    apply_row_result(idx, res)
                except Exception as exc:
                    print(f"  [Worker Error at row {row} | {idx}/{total_rows}]: {exc}")
                    total_failed += len(miss)
    else:
        for idx, (row, original, missing) in enumerate(rows_to_translate, 1):
            print(f"  [Row {row} | {idx}/{total_rows}] {original[:50]}  -> {len(missing)} langs")
            res = _process_row_translation(row, original, missing)
            apply_row_result(idx, res)

    # ── Phase 3: SUMMARY ─────────────────────────────────────────────
    print()
    print("=" * 60)
    ver = bump_version(wb)
    save_workbook(wb)
    print(f"  DONE! Translated: {total_translated}, Failed: {total_failed}")
    print(f"  Version: {ver}")
    print("=" * 60)




# ── validate ─────────────────────────────────────────────────────────

def cmd_validate() -> None:
    """Validate all translations in the dictionary using advanced Validator."""
    from src.core.validator import Validator
    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    # Pass BLOCK_JSON to validator for regex-aware checks
    validator = Validator(str(BLOCK_JSON) if BLOCK_JSON.exists() else None)

    # Build shadok exclusion set by orig only (never by new). Integrity uses strict resolve.
    shadok_row_set: set[int] = set()
    shadok_config = load_shadok_config()
    shadok_resolve_error: str | None = None
    shadok_resolved: list[tuple[int, str, str]] = []
    if shadok_config:
        mapping_for_exclude = shadok_config.get("mapping", [])
        shadok_row_set = build_shadok_exclusion_rows(ws, col_map, mapping_for_exclude)
        try:
            shadok_resolved = resolve_shadok_mapping_rows(ws, col_map, mapping_for_exclude)
        except ValueError as e:
            shadok_resolve_error = str(e)
        if shadok_row_set:
            print(f"  Excluding {len(shadok_row_set)} shadok rows from structural validation.")

    errors = 0
    checked = 0

    print(f"Starting validation for {len([l for l in langs if l != 'ru'])} languages...")

    for row in range(2, ws.max_row + 1):
        if row in shadok_row_set:
            continue  # skip shadok rows
        original = ws.cell(row, col_map["Original"]).value
        if not original:
            continue
        original_str = str(original)
        
        for lc in langs:
            if lc not in col_map or lc == "ru":
                continue
            translation = ws.cell(row, col_map[lc]).value
            if not translation:
                continue
            
            checked += 1
            row_errors = validator.validate_row(original_str, str(translation), lc)
            if row_errors:
                errors += len(row_errors)
                print(f"  [Row {row}][{lc}] Error(s):")
                for err in row_errors:
                    print(f"    - {err}")

    print(f"\n--- Translation checks: {checked} checked, {errors} issues ---")

    # Phase 2: Exact block validation (blocks.json vs Original column)
    if validator.blocks:
        print("\nRunning exact block validation (blocks.json)...")
        originals = {}
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, col_map["Original"]).value
            if val:
                originals[row] = str(val)

        block_errors = validator.validate_blocks(originals)
        if block_errors:
            errors += len(block_errors)
            for err in block_errors:
                print(f"  {err}")
            print(f"--- Block regex checks: {len(block_errors)} issues ---")
        else:
            print("--- Block regex checks: all OK ---")

    # Phase 3: Shadok integrity (resolve + non-empty + max_line_length)
    if shadok_config:
        print("\nRunning Shadok integrity checks...")
        shadok_errors = 0
        mapping = shadok_config.get("mapping", [])
        max_line_len = int(shadok_config.get("max_line_length", 39))
        target_langs = get_shadok_target_langs()

        if shadok_resolve_error:
            shadok_errors += 1
            print(f"  [SHADOK] Resolution failed: {shadok_resolve_error}")
        else:
            if len(shadok_resolved) != len(mapping):
                shadok_errors += 1
                print(
                    f"  [SHADOK] Expected {len(mapping)} mapped rows, resolved {len(shadok_resolved)}"
                )
            for row_idx, orig, _new in shadok_resolved:
                for lc in target_langs:
                    if lc not in col_map:
                        continue
                    cell_val = ws.cell(row_idx, col_map[lc]).value
                    text = "" if cell_val is None else str(cell_val)
                    if not text.strip():
                        shadok_errors += 1
                        print(f"  [SHADOK][Row {row_idx}][{lc}] empty cell")
                        continue
                    vl = visual_length(text)
                    if vl > max_line_len:
                        shadok_errors += 1
                        print(
                            f"  [SHADOK][Row {row_idx}][{lc}] visual_length {vl} > {max_line_len}"
                        )

        errors += shadok_errors
        print(f"--- Shadok integrity: {shadok_errors} issues ---")

    print(f"\nValidation complete. Total issues: {errors}")


# ── export ───────────────────────────────────────────────────────────

def cmd_export() -> None:
    """Export per-language CSV files (original, translation) for the bin builder."""
    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    TRANSLATIONS_DIR.mkdir(parents=True, exist_ok=True)

    missing_total = 0
    
    for lc in langs:
        if lc not in col_map or lc == "ru":
            continue
        csv_path = TRANSLATIONS_DIR / f"{lc}.csv"
        count = 0
        missing_count = 0
        with csv_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["original", "translation"])
            for row in range(2, ws.max_row + 1):
                original = ws.cell(row, col_map["Original"]).value
                if not original:
                    continue
                    
                translation = ws.cell(row, col_map[lc]).value
                
                if not translation or not str(translation).strip():
                    english_fallback = ws.cell(row, col_map["en"]).value if "en" in col_map else None
                    translation = english_fallback if english_fallback and str(english_fallback).strip() else original
                    missing_count += 1
                    missing_total += 1
                    print(f"  [WARNING] Row {row} missing translation for '{lc}'. Fallback to: {translation[:30] + '...' if len(str(translation)) > 30 else translation}")

                writer.writerow([detokenize(str(original)), detokenize(str(translation))])
                count += 1
                
        print(f"  {lc}.csv: {count} entries" + (f" ({missing_count} missing translations filled with fallback)" if missing_count else ""))

    if missing_total > 0:
        print(f"\n[ALERT] Export finished with {missing_total} missing translations!")
        print("Run `python -m src.main translate` to translate missing lines.")
    else:
        print("\nExport done.")

    print("\n[BUILD] Auto-building binaries...")
    cmd_build()


def cmd_sync() -> None:
    """Sync Excel dictionary with translations/ua.csv and ru.csv."""
    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    # Ensure header row
    expected_cols = ["Original"] + list(langs.keys())
    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        for ci, col in enumerate(expected_cols, 1):
            ws.cell(1, ci, col)

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Add missing language columns
    for lang_code in langs:
        if lang_code not in header:
            idx = len(header) + 1
            ws.cell(1, idx, lang_code)
            header.append(lang_code)

    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    # 1. Collect all current keys from Excel
    excel_keys = {}  # original_text -> list of row_indices
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_map["Original"]).value
        if val:
            v_str = str(val)
            if v_str not in excel_keys:
                excel_keys[v_str] = []
            excel_keys[v_str].append(row)

    # 2. Collect all valid keys from source CSVs
    source_data = {} # key -> tokenized_text

    # Read from DATA_DIR (Source of truth), NOT translations/
    for csv_file in [DATA_DIR / "ua.csv", DATA_DIR / "ru.csv"]:
        path = Path(csv_file)
        if not path.exists(): continue

        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) >= 1:
                    orig_tok = tokenize(row[0])
                    if not orig_tok: continue
                    if orig_tok not in source_data:
                        trans = tokenize(row[1]) if len(row) > 1 else orig_tok
                        source_data[orig_tok] = trans

    # 3. Synchronize: Add missing, update existing
    # NOTE: We do NOT delete rows that are not in source CSVs anymore
    # This preserves manually added rows in the dictionary
    added = 0

    for orig_tok, trans_tok in source_data.items():
        if orig_tok not in excel_keys:
            # Add new row
            new_row = ws.max_row + 1
            ws.cell(new_row, col_map["Original"], orig_tok)
            if "ru" in col_map: ws.cell(new_row, col_map["ru"], trans_tok)
            if "ua" in col_map: ws.cell(new_row, col_map["ua"], trans_tok)
            added += 1
            excel_keys[orig_tok] = [new_row]

    # 4. Remove duplicates only (keep first occurrence)
    duplicates_removed = 0
    all_duplicate_rows: list[int] = []
    for orig, row_indices in excel_keys.items():
        if len(row_indices) > 1:
            print(f"  [SYNC] Merging and removing {len(row_indices)-1} duplicate(s) for: {repr(orig[:30])}...")
            first_row = row_indices[0]
            for dup_row in row_indices[1:]:
                for col in range(1, ws.max_column + 1):
                    first_val = ws.cell(first_row, col).value
                    dup_val = ws.cell(dup_row, col).value
                    first_empty = first_val is None or not str(first_val).strip()
                    dup_has_val = dup_val is not None and str(dup_val).strip() != ""
                    if first_empty and dup_has_val:
                        ws.cell(first_row, col, dup_val)
                all_duplicate_rows.append(dup_row)

    for row_idx in sorted(all_duplicate_rows, reverse=True):
        ws.delete_rows(row_idx)
        duplicates_removed += 1

    # Bump version or just update status
    ver = get_version(wb)
    print(f"Sync done. Added: {added}, Duplicates removed: {duplicates_removed}, Version: {ver}")
    save_workbook(wb)


# ── build ────────────────────────────────────────────────────────────

def cmd_build() -> None:
    """Run build_translation_bin.py for each exported CSV."""
    langs = load_languages()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for lc in langs:
        if lc == "ru":
            continue
        csv_path = TRANSLATIONS_DIR / f"{lc}.csv"
        if not csv_path.exists():
            print(f"  Skip {lc}: no CSV")
            continue
        bin_path = OUTPUT_DIR / f"translation_{lc}.bin"
        cmd = [sys.executable, str(BUILD_SCRIPT), str(csv_path), "-o", str(bin_path)]
        print(f"  Building {bin_path.name}...")
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"  [!] Error building {lc}: {result.stderr}")
        else:
            print(f"  OK: {bin_path.name}")

    print("Build done.")


def cmd_dist() -> None:
    """Organize NRO and translation bins into per-language folders in 'dist'."""
    langs = load_languages()

    if DIST_DIR.exists():
        for attempt in range(5):
            try:
                shutil.rmtree(DIST_DIR)
                break
            except PermissionError:
                if attempt < 4:
                    print(f"  [!] dist/ is locked. Retrying in 2s... ({attempt + 1}/5)")
                    time.sleep(2)
                else:
                    print("  [CRITICAL] Could not clean dist/ — close Explorer or other programs using it.")
                    raise
    DIST_DIR.mkdir(parents=True)

    # Find the latest patched DBI NRO file
    source_nro = get_patched_nro_path()
    if not source_nro:
        print("  [ERROR] No DBI.*_patched.nro file found in root!")
        return

    nro_ver = get_nro_version()
    print(f"  Using patched NRO: {source_nro.name} (version {nro_ver})")

    for lc in langs:
        if lc == "ru": continue

        bin_path = OUTPUT_DIR / f"translation_{lc}.bin"
        if not bin_path.exists():
            # Try to build it if missing? No, user usually runs build before dist.
            continue

        lang_dist = DIST_DIR / lc
        lang_dist.mkdir(parents=True, exist_ok=True)

        # Copy and rename NRO to DBI.nro as requested by user's example
        shutil.copy2(source_nro, lang_dist / "DBI.nro")
        # Copy and rename BIN to translation.bin
        shutil.copy2(bin_path, lang_dist / "translation.bin")

        print(f"  [OK] {lc}: DBI.nro + translation.bin")

    print(f"\nOrganization in 'dist' folder complete using {source_nro.name}")


def cmd_align() -> None:
    """Align colons in blocks by longest line per language per block."""

    blocks = load_blocks()
    if not blocks:
        print("No blocks defined in blocks.json.")
        return

    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    # Explicitly exclude Shadok rows from alignment mutations
    shadok_row_set: set[int] = set()
    shadok_config = load_shadok_config()
    if shadok_config:
        shadok_row_set = build_shadok_exclusion_rows(
            ws, col_map, shadok_config.get("mapping", [])
        )
        if shadok_row_set:
            print(f"  Excluding {len(shadok_row_set)} shadok rows from align.")

    # Cache all Original values for faster matching
    originals = {}  # row -> original_str
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_map["Original"]).value
        if val:
            originals[row] = str(val)

    # Build grouped_data[block_id] = list of matched row indices
    grouped_data = {}  # block_id -> [row_idx, ...]

    for bid, patterns in blocks.items():
        matched_rows = []
        for pattern in patterns:
            found = False
            skipped_shadok = False
            for row, orig_val in originals.items():
                if orig_val == pattern:
                    if row in shadok_row_set:
                        skipped_shadok = True
                        continue
                    matched_rows.append(row)
                    found = True
                    break

            if not found and not skipped_shadok:
                print(f"  [WARN] String not found for block {bid}: {pattern[:50]}...")

        matched_rows = [r for r in matched_rows if r not in shadok_row_set]
        if matched_rows:
            grouped_data[bid] = matched_rows

    if not grouped_data:
        print("No rows found matching blocks.json patterns.")
        return

    lang_cols = {lc: col_map[lc] for lc in langs if lc in col_map and lc != "ru"}

    affected_count = 0

    # Process each block
    for bid, matched_rows in grouped_data.items():
        print(f"  Aligning block: {bid:<25} (Rows: {len(matched_rows)})")

        # Special handling for NSP_INSTALL_ANSWERS: pad to exactly 4 chars
        if bid == "NSP_INSTALL_ANSWERS":
            for lc, col_idx in lang_cols.items():
                for row_idx in matched_rows:
                    val = ws.cell(row_idx, col_idx).value
                    if not val:
                        continue
                    val_str = str(val)
                    stripped = val_str.strip()
                    char_count = len(stripped)

                    if char_count >= 4:
                        new_val = stripped[:4]
                    elif char_count == 3:
                        new_val = stripped + " "
                    elif char_count == 2:
                        new_val = " " + stripped + " "
                    elif char_count == 1:
                        new_val = " " + stripped + "  "
                    else:
                        new_val = "    "

                    if new_val != val_str:
                        ws.cell(row_idx, col_idx, new_val)
                        affected_count += 1
            continue

        # Special handling for TITLE_INFO block with double-colon strings
        is_title_info = (bid == "TITLE_INFO")

        # For each language, find max length and align
        for lc, col_idx in lang_cols.items():
            # Find max prefix length for this language in this block
            max_len = 0
            for row_idx in matched_rows:
                val = ws.cell(row_idx, col_idx).value
                if not val:
                    continue
                val_str = str(val).strip()

                if ":" in val_str:
                    # For TITLE_INFO: only consider first colon for alignment
                    if is_title_info:
                        prefix = val_str.split(":", 1)[0]
                    else:
                        prefix = val_str.split(":", 1)[0]
                    clean_prefix = prefix.rstrip()
                    max_len = max(max_len, len(clean_prefix))

            if max_len == 0:
                continue

            # Align all rows in this block for this language
            target_len = max_len + 1  # +1 for at least one space before colon

            for row_idx in matched_rows:
                val = ws.cell(row_idx, col_idx).value
                if not val:
                    continue
                val_str = str(val).strip()

                if ":" in val_str:
                    prefix, suffix = val_str.split(":", 1)
                    clean_prefix = prefix.rstrip()
                    current_len = len(clean_prefix)

                    padding = target_len - current_len
                    if padding < 1:
                        padding = 1

                    new_val = clean_prefix + (' ' * padding) + ':' + suffix

                    if new_val != val_str:
                        ws.cell(row_idx, col_idx, new_val)
                        affected_count += 1

    if affected_count > 0:
        ver = bump_version(wb)
        save_workbook(wb)
        print(f"\nAlignment done. Adjusted {affected_count} cells. Version: {ver}")
    else:
        print("\nAlignment done. No changes needed.")




# ── shadok ───────────────────────────────────────────────────────────

def cmd_shadok() -> None:
    """Localize Shadok parody texts (mapping.new) into rows identified by mapping.orig."""
    config = load_shadok_config()
    if not config:
        print("  [ERROR] data/shadok.json not found.")
        return

    mapping = config.get("mapping", [])
    if not mapping:
        print("  [ERROR] shadok.json mapping is empty.")
        return

    max_line_length = int(config.get("max_line_length", 39))
    target_langs = get_shadok_target_langs()

    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    try:
        resolved = resolve_shadok_mapping_rows(ws, col_map, mapping)
    except ValueError as e:
        print(f"  [ERROR] {e}")
        print("  No cells written.")
        return

    source = "\n".join(new for _, _, new in resolved)
    expected_count = len(resolved)

    print()
    print("=" * 60)
    print("  DBI SHADOK LOCALIZER")
    print("=" * 60)
    print(f"  Provider  : {ai_client.PROVIDER} (model: {ai_client.MODEL})")
    print(f"  Rows      : {expected_count}")
    print(f"  Max len   : {max_line_length}")
    print(f"  Languages : {len(target_langs)} ({', '.join(target_langs)})")
    print("-" * 60)

    init_session_shadok()

    ok_langs = 0
    fail_langs = 0

    for lc in target_langs:
        if lc not in col_map:
            print(f"  [SKIP][{lc}] no workbook column")
            continue

        print(f"  [SHADOK] -> {lc} ...")
        try:
            result = translate_shadok_block(
                source,
                [lc],
                max_line_length,
                expected_lines=expected_count,
            )
            if lc not in result:
                raise ValueError(f"Missing language key {lc!r} in AI response")
            lines = parse_and_validate_shadok_block(
                result[lc], expected_count, max_line_length
            )
        except Exception as e:
            print(f"  [ERROR][{lc}] {e}")
            print(f"  [ERROR][{lc}] Writing zero cells for this language.")
            fail_langs += 1
            continue

        for (row_idx, _orig, _new), line in zip(resolved, lines):
            ws.cell(row_idx, col_map[lc], line)
        save_workbook(wb)
        ok_langs += 1
        print(f"  [OK][{lc}] Wrote {expected_count} cells")

    print()
    print("=" * 60)
    print(f"  DONE! OK languages: {ok_langs}, Failed: {fail_langs}")
    print("  Original column unchanged. No version bump.")
    print("=" * 60)


# ── clear ────────────────────────────────────────────────────────────

def cmd_clear(lang_code: str) -> None:
    """Clear all translations for a specific language in the dictionary."""
    if not DICT_PATH.exists():
        print("No dictionary found.")
        return

    wb = openpyxl.load_workbook(DICT_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet {SHEET_NAME} not found.")
        return

    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    if lang_code not in col_map:
        print(f"Language '{lang_code}' not found in dictionary columns.")
        return

    col_idx = col_map[lang_code]
    cleared = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, col_idx).value:
            ws.cell(row, col_idx).value = None
            cleared += 1

    ver = bump_version(wb)
    save_workbook(wb)
    print(f"Cleared {cleared} entries for '{lang_code}'. Version: {ver}")


# ── main ─────────────────────────────────────────────────────────────

def cmd_deploy() -> None:
    """Commit, push and create a GitHub release with assets."""

    # 1. Get versions & Check completeness
    try:
        patched_nro = get_patched_nro_path()
        if not patched_nro:
            print("  [ERROR] No DBI.*_patched.nro file found!")
            return

        dbi_ver = get_nro_version()
        print(f"  Using patched NRO: {patched_nro.name} (version {dbi_ver})")

        wb = open_or_create_workbook()
        ws = wb[SHEET_NAME]
        patcher_ver = get_version(wb)

        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {h: i + 1 for i, h in enumerate(header) if h}
        langs = load_languages()
        lang_codes = [lc for lc in langs if lc in col_map and lc != "ru"]

        print("  [CHECK] Verifying translation completeness...")
        missing_count = 0
        for row in range(2, ws.max_row + 1):
            if not ws.cell(row, col_map["Original"]).value: continue
            for lc in lang_codes:
                val = ws.cell(row, col_map[lc]).value
                if not val or not str(val).strip():
                    missing_count += 1

        if missing_count > 0:
            print(f"  [ERROR] Cannot deploy: Found {missing_count} missing translations!")
            print("  Please run 'python -m src.main translate' first.")
            return

    except Exception as e:
        print(f"  [ERROR] Preparation failed: {e}")
        return

    # 2. Copy files to target directories (always execute)
    print("  [COPY] Copying files to target directories...")
    try:
        # Find the patched NRO
        patched_nro = get_patched_nro_path()

        if not patched_nro:
            print(f"  [WARN] No patched NRO found for version {dbi_ver}")
        else:
            print(f"  [COPY] Using: {patched_nro.name}")
            # Copy to D:\git\dev\_kefir\kefir\switch\DBI\DBI.nro
            kefir_dir = Path("D:/git/dev/_kefir/kefir/switch/DBI")
            kefir_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(patched_nro, kefir_dir / "DBI.nro")
            print(f"  [COPY] {patched_nro.name} -> {kefir_dir / 'DBI.nro'}")

            # Copy translation_en.bin to D:\git\dev\_kefir\kefir\switch\DBI\translation.bin
            en_bin = OUTPUT_DIR / "translation_en.bin"
            if en_bin.exists():
                shutil.copy2(en_bin, kefir_dir / "translation.bin")
                print(f"  [COPY] translation_en.bin -> {kefir_dir / 'translation.bin'}")
            else:
                print(f"  [WARN] translation_en.bin not found")

            # Copy translation_ua.bin to E:\Switch\addons\switch\DBI\translation.bin
            ua_bin = OUTPUT_DIR / "translation_ua.bin"
            switch_dir = Path("E:/Switch/addons/switch/DBI")
            if ua_bin.exists():
                switch_dir.mkdir(parents=True, exist_ok=True)
                shutil.copy2(ua_bin, switch_dir / "translation.bin")
                print(f"  [COPY] translation_ua.bin -> {switch_dir / 'translation.bin'}")
            else:
                print(f"  [WARN] translation_ua.bin not found")

        print("  [COPY] File copying completed!")
    except Exception as e:
        print(f"  [ERROR] File copying failed: {e}")

    # 3. Git operations
    print("  [GIT] Staging changes and pushing...")
    try:
        subprocess.run(["git", "add", "."], check=True)
        # Check if there are changes to commit
        status = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True, encoding="utf-8").stdout
        if status.strip():
            subprocess.run(["git", "commit", "-m", f"chore: deploy DBI {dbi_ver} localization (v{patcher_ver})"], check=True)
            subprocess.run(["git", "push", "origin", "master"], check=True)
            print("  [GIT] Changes pushed successfully.")
        else:
            print("  [GIT] No changes to commit.")
    except subprocess.CalledProcessError as e:
        print(f"  [ERROR] Git operation failed: {e}")
        return

    # 4. Prepare release body
    # (template remains same)
    langs_list = """*   **BE** — Belarusian
*   **DE** — German
*   **EN** — English (US)
*   **ENGB** — English (UK)
*   **ES** — Spanish (Spain)
*   **ES419** — Spanish (Latin America)
*   **ET** — Estonian
*   **FR** — French
*   **FRCA** — French (Canada)
*   **IT** — Italian
*   **JP** — Japanese
*   **KK** — Kazakh
*   **KR** — Korean
*   **LT** — Lithuanian
*   **LV** — Latvian
*   **NL** — Dutch
*   **PL** — Polish
*   **PT** — Portuguese (Portugal)
*   **PTBR** — Portuguese (Brazil)
*   **UA** — Ukrainian
*   **ZHCN** — Chinese (Simplified)
*   **ZHTW** — Chinese (Traditional)
*   **TR** — Turkish"""

    release_body = f"""### 🌍 DBI Multilingual Localization (v{patcher_ver})
![GitHub release (tag)](https://img.shields.io/github/downloads/rashevskyv/DBIPatcher/{dbi_ver}/total)

This release provides high-quality translations for **DBI version {dbi_ver}**.

> [!IMPORTANT]
> This translation is **strictly compatible only with the DBI.nro version provided in this release**. Do not use it with other versions of DBI as it may cause UI glitches or crashes.

### 📦 Supported Languages
{langs_list}

***

### 🛠️ Installation Instruction
1. Download **`DBI.nro`** (Patched/Compatible version) from this release.
2. Download the **`translation_XX.bin`** file for your desired language.
3. **Rename** the translation file to exactly `translation.bin`.
4. Place both `DBI.nro` and `translation.bin` into the `/switch/DBI/` folder on your SD card.

### ⚠️ Known Issues
- ~~**Hardcoded Strings**: Some interface elements are hardcoded within the DBI binary and cannot be localized via `translation.bin`. Confirmation prompts may still display in Russian: **Да** (Yes) and **Нет** (No).~~ ✅ Fixed!
- **Shadok Fables**: Satirical blocks use intentional adapted parody localization (not literal DBI Shadok translation) via `python -m src.main shadok`.
- ~~**System Language Names**: Names of languages in the DBI settings menu are hardcoded in the binary.~~ ✅ Fixed!
- **Launcher Compatibility**: Translations have been tested exclusively on [Kefir](https://github.com/rashevskyv/kefir). On Kefir, the translation works successfully regardless of whether DBI is launched via [Sphaira](https://github.com/ITotalJustice/sphaira) or [nx-hbmenu](https://github.com/switchbrew/nx-hbmenu/releases/). If you experience issues with translations not applying on other custom firmwares, please refer to [#12](https://github.com/rashevskyv/DBIPatcher/issues/12).

### 📝 Translation Coverage
This community translation set is still evolving. Some strings may remain untranslated and appear in the original DBI language or use an English fallback. Please report missing or incorrect strings in the project issues.

### 🙏 Credits
- [Bohdan Buinich](https://github.com/BohdanBuinich) — creator of the earlier `dbi-i18n` runtime-translation approach and compatible table format.
- [0xroast](https://github.com/0xroast) — author of `dbi-translate`, the pinned DBI 905 runtime patcher used for this release.

***
*Note: This NRO is a modified version of the [original DBI](https://github.com/rashevskyv/dbi/releases/tag/{dbi_ver}ru) optimized for these translations.*
"""
    
    body_path = Path("scratch/release_body.md")
    body_path.parent.mkdir(exist_ok=True)
    body_path.write_text(release_body, encoding="utf-8")

    # 5. GitHub Release - Create or update
    print(f"  [GH] Checking if release {dbi_ver} exists...")
    check_tag = subprocess.run(["gh", "release", "view", dbi_ver], capture_output=True, text=True, encoding="utf-8")

    assets = [str(a) for a in Path("output").glob("translation_*.bin")]

    # Get the patched NRO for release and rename it for the asset upload
    patched_nro = get_patched_nro_path()
    nro_assets = []
    if patched_nro:
        print(f"  [GH] Preparing NRO asset: {patched_nro.name} -> DBI.nro")
        release_nro = Path("scratch/DBI.nro")
        shutil.copy2(patched_nro, release_nro)
        nro_assets = [str(release_nro)]
    else:
        print(f"  [WARN] No patched NRO found to include in release")

    if check_tag.returncode == 0:
        # Release exists - update assets and notes
        print(f"  [GH] Release {dbi_ver} exists, updating assets and notes...")

        # Add update notice to release body
        from datetime import datetime
        from zoneinfo import ZoneInfo
        kyiv_tz = ZoneInfo("Europe/Kyiv")
        kyiv_time = datetime.now(kyiv_tz).strftime("%Y-%m-%d %H:%M")

        update_notice = f"""> [!WARNING]
> 🔄 **Release updated on {kyiv_time} (Kyiv time).** Please re-download both **DBI.nro** and **translation files** to get the latest version.
"""
        # Insert update notice after the badge line
        badge_line = f"![GitHub release (tag)](https://img.shields.io/github/downloads/rashevskyv/DBIPatcher/{dbi_ver}/total)"
        release_body = release_body.replace(
            badge_line,
            badge_line + "\n\n" + update_notice
        )
        body_path.write_text(release_body, encoding="utf-8")

        # Update release notes
        try:
            subprocess.run(["gh", "release", "edit", dbi_ver, "--notes-file", str(body_path)], check=True)
            print(f"  [GH] Release notes updated for {dbi_ver}")
        except subprocess.CalledProcessError as e:
            print(f"  [ERROR] Failed to update release notes: {e}")

        # Upload all assets with --clobber to overwrite existing files
        upload_cmd = ["gh", "release", "upload", dbi_ver, "--clobber"]
        upload_cmd.extend(assets)
        upload_cmd.extend(nro_assets)

        try:
            subprocess.run(upload_cmd, check=True)
            print(f"  [GH] Assets updated successfully in release {dbi_ver}!")
        except subprocess.CalledProcessError as e:
            print(f"  [ERROR] Failed to update assets: {e}")
    else:
        # Release doesn't exist - create new one
        print(f"  [GH] Creating new release {dbi_ver}...")

        cmd = [
            "gh", "release", "create", dbi_ver,
            "--title", f"DBI {dbi_ver} Localization",
            "--notes-file", str(body_path)
        ]
        cmd.extend(assets)
        cmd.extend(nro_assets)

        try:
            subprocess.run(cmd, check=True)
            print(f"  [GH] Release {dbi_ver} created successfully!")
        except subprocess.CalledProcessError as e:
            print(f"  [ERROR] GitHub release failed: {e}")


def cmd_check() -> None:
    """Check dictionary integrity against source CSV and blocks.json"""
    print("\n" + "="*60 + "\n  STEP: check\n" + "="*60)
    
    ua_path = DATA_DIR / "ua.csv"
    ua_originals = []
    if ua_path.exists():
        with open(ua_path, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if row:
                    ua_originals.append(tokenize(row[0]))
                    
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}
    
    dict_originals = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, col_map["Original"]).value
        if v:
            dict_originals.add(str(v))
            
    mismatches = 0
    for k in ua_originals:
        if k not in dict_originals:
            print(f"  [ERROR] Source key missing or altered in dictionary: {repr(k)}")
            mismatches += 1
            
    blocks = load_blocks()
    if blocks:
        for bid, patterns in blocks.items():
            for pat in patterns:
                if pat not in dict_originals:
                    print(f"  [ERROR] Block {bid} string missing in dictionary: {repr(pat)}")
                    mismatches += 1
                
    if mismatches == 0:
        print("  [OK] Health check passed successfully! All strings match identically.")
    else:
        print(f"\n  [FAIL] Health check failed with {mismatches} issues.")


COMMANDS = {
    "sync": cmd_sync,
    "translate": cmd_translate,
    "shadok": cmd_shadok,
    "validate": cmd_validate,
    "align": cmd_align,
    "export": cmd_export,
    "build": cmd_build,
    "dist": cmd_dist,
    "clear": cmd_clear,
    "deploy": cmd_deploy,
    "check": cmd_check,
    "test": lambda: cmd_test(),
}


def cmd_all() -> None:
    for name in ("sync", "translate", "align", "validate", "export", "build", "dist"):
        print(f"\n{'='*60}\n  STEP: {name}\n{'='*60}")
        COMMANDS[name]()


def cmd_test() -> None:
    """Run all steps except deploy (sync, translate, align, validate, export, build, dist)."""
    for name in ("sync", "translate", "align", "validate", "export", "build", "dist"):
        print(f"\n{'='*60}\n  STEP: {name}\n{'='*60}")
        COMMANDS[name]()
    print(f"\n{'='*60}\n  TEST COMPLETE\n{'='*60}")
    print("All steps completed successfully. Ready for deployment.")
    print("To deploy, run: python -m src.main deploy")


def cmd_help() -> None:
    """Display all available commands with descriptions."""
    print("\n" + "="*60)
    print("  DBI TRANSLATION PIPELINE - Available Commands")
    print("="*60)
    print("\nUsage: python -m src.main <command> [command2 ...] [options]\n")
    print("Commands:")
    print("  sync        - Sync ua.csv into dictionary.xlsx")
    print("  translate   - Translate missing cells via AI")
    print("  shadok      - Localize Shadok parody block via AI (manual)")
    print("  validate    - Validate all translations")
    print("  align       - Align colons in blocks by longest line")
    print("  export      - Export per-language CSVs")
    print("  build       - Build .bin files from CSVs")
    print("  dist        - Organize NRO and bins into dist folders")
    print("  clear       - Clear all translations for a language")
    print("                Usage: python -m src.main clear <lang_code>")
    print("  deploy      - Commit, push and create GitHub release")
    print("  check       - Check dictionary integrity")
    print("  test        - Run all steps except deploy")
    print("  all         - Run full pipeline (sync → dist)")
    print("  help        - Show this help message")
    print("\nOptions:")
    print("  -f, --force - Force re-translate all strings")
    print("\nExamples:")
    print("  python -m src.main sync")
    print("  python -m src.main translate -f")
    print("  python -m src.main shadok")
    print("  python -m src.main align build")
    print("  python -m src.main export build dist")
    print("  python -m src.main clear ua")
    print("  python -m src.main all")
    print("="*60 + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="DBI Translation Pipeline", add_help=False)
    parser.add_argument("commands", nargs="*", help="Pipeline steps to run")
    parser.add_argument("-f", "--force", action="store_true", help="Force re-translate all strings")
    parser.add_argument("-h", "--help", action="store_true", help="Show help message")
    args = parser.parse_args()

    if args.help or not args.commands or (len(args.commands) == 1 and args.commands[0] == "help"):
        cmd_help()
        return 0

    # Validate commands
    valid_commands = {*COMMANDS.keys(), "all", "help"}
    for cmd in args.commands:
        if cmd not in valid_commands and cmd not in load_languages():
            print(f"Error: Unknown command '{cmd}'")
            print(f"Valid commands: {', '.join(sorted(valid_commands))}")
            return 1

    # Clear log at every start
    log_path = ROOT / "logs" / "ai_proxy.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    from datetime import datetime, timezone
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"--- SESSION: {' '.join(args.commands)} | {datetime.now(timezone.utc).isoformat()} ---\n")

    # Execute commands sequentially
    for i, cmd in enumerate(args.commands):
        if len(args.commands) > 1:
            print(f"\n{'='*60}\n  STEP {i+1}/{len(args.commands)}: {cmd}\n{'='*60}")

        if cmd == "all":
            cmd_all()
        elif cmd == "clear":
            # For clear command, next argument should be language code
            if i + 1 >= len(args.commands):
                print("Error: 'clear' command requires a language code (e.g., 'ua')")
                return 1
            lang_code = args.commands[i + 1]
            cmd_clear(lang_code)
            # Skip next argument as it was the language code
            args.commands[i + 1] = None
        elif cmd is None:
            # Skip (was consumed as language code for clear)
            continue
        else:
            COMMANDS[cmd]()

    return 0


if __name__ == "__main__":
    sys.exit(main())
