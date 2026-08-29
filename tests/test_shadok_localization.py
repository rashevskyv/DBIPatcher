"""Offline tests for safe Shadok localization (cmd_shadok / translate / align / validate)."""

from __future__ import annotations

import json
import sys
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.core import ai_client  # noqa: E402
from src.core.ai_client import (  # noqa: E402
    _parse_json_safe,
    build_shadok_system_prompt,
)
from src.main import (  # noqa: E402
    COMMANDS,
    cmd_align,
    cmd_shadok,
    cmd_translate,
    cmd_validate,
    fit_shadok_lines_to_slots,
    get_shadok_target_langs,
    load_shadok_config,
    pad_shadok_lines_to_mapping,
    parse_and_validate_shadok_block,
    resolve_shadok_mapping_rows,
)


def _load_mapping() -> list[dict]:
    config = load_shadok_config()
    assert config is not None
    return list(config["mapping"])


def _make_workbook(
    mapping: list[dict],
    lang_codes: list[str],
    *,
    seed_lang_values: dict[str, str] | None = None,
    extra_rows: list[tuple[str, dict[str, str]]] | None = None,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"
    meta = wb.create_sheet("Metadata")
    meta["A1"] = "version"
    meta["B1"] = "0.0.87"
    meta["A2"] = "updated"
    meta["B2"] = ""

    headers = ["Original"] + lang_codes
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)

    seed_lang_values = seed_lang_values or {}
    for i, item in enumerate(mapping):
        row = i + 2
        ws.cell(row, 1, item["orig"])
        for col, lc in enumerate(lang_codes, 2):
            if lc in seed_lang_values:
                ws.cell(row, col, seed_lang_values[lc])

    if extra_rows:
        start = len(mapping) + 2
        for offset, (orig, langs) in enumerate(extra_rows):
            row = start + offset
            ws.cell(row, 1, orig)
            for col, lc in enumerate(lang_codes, 2):
                if lc in langs:
                    ws.cell(row, col, langs[lc])
    return wb


def _col_map(ws) -> dict[str, int]:
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(header) if h}


def _valid_block(prefix: str, count: int = 33) -> str:
    # Keep well under visual_length 39
    return "\n".join(f"{prefix}{i:02d} ok line" for i in range(count))


class ShadokLocalizationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.mapping = _load_mapping()
        self.assertEqual(len(self.mapping), 33)
        self.lang_codes = ["en", "frca", "tr", "de"]
        self.langs = {
            "en": "English",
            "frca": "French (Canada)",
            "tr": "Turkish",
            "de": "German",
        }

    def test_target_langs_include_frca_and_tr(self) -> None:
        targets = get_shadok_target_langs()
        self.assertIn("frca", targets)
        self.assertIn("tr", targets)
        self.assertNotIn("ru", targets)
        with open(ROOT / "data" / "languages.json", encoding="utf-8") as f:
            all_langs = json.load(f)
        expected = [lc for lc in all_langs if lc != "ru"]
        self.assertEqual(targets, expected)

    def test_shadok_prompt_requires_screen_reflow(self) -> None:
        with open(ROOT / "data" / "prompts.json", encoding="utf-8") as f:
            prompt = json.load(f)["shadok"]
        self.assertIn("WORD-WRAP / REFLOW", prompt)
        self.assertIn("max_line_length", prompt)
        self.assertIn("expected_lines", prompt)
        self.assertIn("Russian-Ukrainian war", prompt)
        self.assertIn("swamp", prompt.lower())
        self.assertNotIn("Do NOT merge or split lines", prompt)
        self.assertNotIn("Do NOT care about line breaks", prompt)

    def test_parse_and_validate_rejects_malformed(self) -> None:
        # Screen height budget is 35; width is 39
        good = _valid_block("L", 33)
        self.assertEqual(len(parse_and_validate_shadok_block(good, 35, 39)), 33)
        self.assertEqual(len(parse_and_validate_shadok_block(_valid_block("L", 30), 35, 39)), 30)
        self.assertEqual(len(parse_and_validate_shadok_block(_valid_block("L", 34), 35, 39)), 34)
        self.assertEqual(len(parse_and_validate_shadok_block(_valid_block("L", 35), 35, 39)), 35)

        with self.assertRaises(ValueError):
            parse_and_validate_shadok_block(_valid_block("L", 36), 35, 39)  # height overflow
        with self.assertRaises(ValueError):
            bad_empty = "\n".join(["ok"] * 20 + ["   "])
            parse_and_validate_shadok_block(bad_empty, 35, 39)
        with self.assertRaises(ValueError):
            too_long = "\n".join(["x" * 40] + ["ok"] * 20)
            parse_and_validate_shadok_block(too_long, 35, 39)

    def test_pad_shadok_lines_uses_space_tail(self) -> None:
        padded = pad_shadok_lines_to_mapping(["a", "b"], 5)
        self.assertEqual(padded, ["a", "b", " ", " ", " "])

    def test_fit_packs_overflow_into_last_slot_with_lf(self) -> None:
        # 35 visual lines, 33 dict slots → last slot holds lines 33..35 via [[LF]]
        lines = [f"L{i}" for i in range(35)]
        fitted = fit_shadok_lines_to_slots(lines, slot_count=33, max_line_length=39)
        self.assertEqual(len(fitted), 33)
        self.assertEqual(fitted[32], "L32[[LF]]L33[[LF]]L34")
        for i in range(32):
            self.assertEqual(fitted[i], f"L{i}")

    def test_resolve_missing_and_duplicate(self) -> None:
        wb = _make_workbook(self.mapping[:2], self.lang_codes)
        ws = wb["Translations"]
        col_map = _col_map(ws)
        with self.assertRaises(ValueError):
            resolve_shadok_mapping_rows(ws, col_map, self.mapping)

        wb2 = _make_workbook(self.mapping[:1], self.lang_codes)
        ws2 = wb2["Translations"]
        ws2.cell(3, 1, self.mapping[0]["orig"])  # duplicate
        col_map2 = _col_map(ws2)
        with self.assertRaises(ValueError):
            resolve_shadok_mapping_rows(ws2, col_map2, self.mapping[:1])

    def test_translate_shadok_block_payload_includes_limits(self) -> None:
        captured = {}

        def fake_retry(url, safe_data, payload, row_id=None, is_shadok=False, is_refine=False):
            captured["payload"] = payload
            captured["safe_data"] = safe_data
            return {"en": _valid_block("E")}

        with patch.object(ai_client, "PROVIDER", "WEB2API"), \
             patch.object(ai_client, "_make_request_with_retry", side_effect=fake_retry):
            ai_client.translate_shadok_block("a\nb\nc", ["en"], 39, expected_lines=35)

        user = json.loads(captured["payload"]["messages"][-1]["content"])
        self.assertEqual(user["max_line_length"], 39)
        self.assertEqual(user["expected_lines"], 35)
        self.assertEqual(user["text"], "a\nb\nc")
        self.assertEqual(user["languages"], ["en"])
        self.assertEqual(user["attempt"], 0)
        system = captured["payload"]["messages"][0]["content"]
        self.assertIn("HARD SCREEN BUDGET", system)
        self.assertEqual(captured["payload"]["model"], "gemini-3.7-flash")

    @patch("src.main.save_workbook")
    @patch("src.main.init_session_shadok")
    @patch("src.main.translate_shadok_block")
    @patch("src.main.open_or_create_workbook")
    def test_cmd_shadok_uses_new_keeps_orig_writes_valid_block(
        self,
        mock_open_wb: MagicMock,
        mock_translate: MagicMock,
        mock_init: MagicMock,
        mock_save: MagicMock,
    ) -> None:
        wb = _make_workbook(
            self.mapping,
            self.lang_codes,
            seed_lang_values={"en": "OLD_EN", "frca": "OLD_FRCA"},
        )
        mock_open_wb.return_value = wb
        ws = wb["Translations"]
        col_map = _col_map(ws)
        originals_before = [
            ws.cell(r, col_map["Original"]).value for r in range(2, 35)
        ]

        def fake_block(
            full_text,
            target_langs,
            max_line_length,
            expected_lines=None,
            attempt=0,
            previous_error=None,
            previous_text=None,
        ):
            self.assertEqual(target_langs, ["en"])
            self.assertEqual(max_line_length, 39)
            self.assertEqual(expected_lines, 35)
            self.assertEqual(attempt, 0)
            # AI must receive joined *new* lines, never orig
            expected_source = "\n".join(item["new"] for item in self.mapping)
            self.assertEqual(full_text, expected_source)
            for item in self.mapping:
                self.assertNotIn(item["orig"], full_text.split("\n"))
            return {"en": _valid_block("EN")}

        mock_translate.side_effect = fake_block

        with patch("src.main.get_shadok_target_langs", return_value=["en"]), \
             patch("sys.argv", ["main.py", "shadok"]):
            cmd_shadok()

        for r in range(2, 35):
            self.assertEqual(ws.cell(r, col_map["Original"]).value, originals_before[r - 2])
            self.assertEqual(ws.cell(r, col_map["en"]).value, f"EN{r - 2:02d} ok line")
            # untouched other language
            self.assertEqual(ws.cell(r, col_map["frca"]).value, "OLD_FRCA")

        self.assertEqual(mock_translate.call_count, 1)
        mock_save.assert_called()
        self.assertIn("shadok", COMMANDS)

    @patch("src.main.save_workbook")
    @patch("src.main.init_session_shadok")
    @patch("src.main.translate_shadok_block")
    @patch("src.main.open_or_create_workbook")
    def test_cmd_shadok_malformed_writes_zero_cells(
        self,
        mock_open_wb: MagicMock,
        mock_translate: MagicMock,
        mock_init: MagicMock,
        mock_save: MagicMock,
    ) -> None:
        wb = _make_workbook(
            self.mapping,
            self.lang_codes,
            seed_lang_values={"en": "KEEP_EN", "tr": "KEEP_TR"},
        )
        mock_open_wb.return_value = wb
        ws = wb["Translations"]
        col_map = _col_map(ws)

        def fake_block(
            full_text,
            target_langs,
            max_line_length,
            expected_lines=None,
            attempt=0,
            previous_error=None,
            previous_text=None,
        ):
            lc = target_langs[0]
            if lc == "en":
                return {"en": _valid_block("BAD", 36)}  # height overflow every attempt
            if lc == "tr":
                return {"tr": "\n".join(["ok"] * 20 + [" "])}  # whitespace-only hole
            return {lc: _valid_block(lc.upper())}

        mock_translate.side_effect = fake_block

        with patch("src.main.get_shadok_target_langs", return_value=["en", "tr"]), \
             patch("sys.argv", ["main.py", "shadok"]):
            cmd_shadok()

        for r in range(2, 35):
            self.assertEqual(ws.cell(r, col_map["en"]).value, "KEEP_EN")
            self.assertEqual(ws.cell(r, col_map["tr"]).value, "KEEP_TR")
            self.assertEqual(ws.cell(r, col_map["Original"]).value, self.mapping[r - 2]["orig"])

        # 3 attempts × 2 langs; never checkpoint-write
        self.assertEqual(mock_translate.call_count, 6)
        mock_save.assert_not_called()

    @patch("src.main.save_workbook")
    @patch("src.main.init_session_shadok")
    @patch("src.main.translate_shadok_block")
    @patch("src.main.open_or_create_workbook")
    def test_cmd_shadok_overlength_writes_zero_cells(
        self,
        mock_open_wb: MagicMock,
        mock_translate: MagicMock,
        mock_init: MagicMock,
        mock_save: MagicMock,
    ) -> None:
        wb = _make_workbook(
            self.mapping,
            self.lang_codes,
            seed_lang_values={"frca": "KEEP_FRCA"},
        )
        mock_open_wb.return_value = wb
        ws = wb["Translations"]
        col_map = _col_map(ws)

        over = "\n".join(["Y" * 40] + [f"ok{i}" for i in range(32)])
        mock_translate.return_value = {"frca": over}

        with patch("src.main.get_shadok_target_langs", return_value=["frca"]), \
             patch("sys.argv", ["main.py", "shadok"]):
            cmd_shadok()

        for r in range(2, 35):
            self.assertEqual(ws.cell(r, col_map["frca"]).value, "KEEP_FRCA")
        self.assertEqual(mock_translate.call_count, 3)
        mock_save.assert_not_called()

    @patch("src.main.save_workbook")
    @patch("src.main.init_session")
    @patch("src.main.open_or_create_workbook")
    @patch("src.main._process_row_translation")
    def test_cmd_translate_skips_shadok_rows(
        self,
        mock_process: MagicMock,
        mock_open_wb: MagicMock,
        mock_init: MagicMock,
        mock_save: MagicMock,
    ) -> None:
        from src.main import RowTranslationResult

        normal_orig = "Обычная строка для перевода"
        wb = _make_workbook(
            self.mapping,
            self.lang_codes,
            extra_rows=[(normal_orig, {})],
        )
        mock_open_wb.return_value = wb

        called_originals: list[str] = []

        def fake_process(row_id, original, missing):
            called_originals.append(str(original))
            return RowTranslationResult(
                row_id=row_id,
                original=original,
                missing=list(missing),
                accepted_translations={lc: "OK" for lc in missing},
            )

        mock_process.side_effect = fake_process

        with patch("src.main.load_languages", return_value=self.langs), \
             patch("src.main.bump_version", return_value="0.0.87"), \
             patch("src.main.get_translate_workers", return_value=1), \
             patch("sys.argv", ["main.py", "translate"]):
            cmd_translate()

        shadok_origs = {item["orig"] for item in self.mapping}
        self.assertTrue(called_originals, "Expected non-shadok worker call")
        for orig in called_originals:
            self.assertNotIn(orig, shadok_origs)
        self.assertIn(normal_orig, called_originals)

        ws = wb["Translations"]
        col_map = _col_map(ws)
        # Shadok en cells remain empty (never scheduled)
        for r in range(2, 35):
            self.assertFalse(str(ws.cell(r, col_map["en"]).value or "").strip())

    @patch("src.main.save_workbook")
    @patch("src.main.open_or_create_workbook")
    def test_cmd_align_does_not_modify_shadok(
        self,
        mock_open_wb: MagicMock,
        mock_save: MagicMock,
    ) -> None:
        shadok_orig = self.mapping[0]["orig"]
        other_orig = "Other Label: value"
        wb = _make_workbook(
            self.mapping,
            self.lang_codes,
            seed_lang_values={"en": "Shadok:short"},
            extra_rows=[(other_orig, {"en": "Other:x"})],
        )
        mock_open_wb.return_value = wb
        ws = wb["Translations"]
        col_map = _col_map(ws)
        shadok_row = resolve_shadok_mapping_rows(ws, col_map, self.mapping)[0][0]
        before = ws.cell(shadok_row, col_map["en"]).value

        fake_blocks = {
            "FAKE_SHADOK": [shadok_orig],
            "FAKE_OTHER": [other_orig],
        }

        with patch("src.main.load_blocks", return_value=fake_blocks), \
             patch("src.main.load_languages", return_value=self.langs), \
             patch("src.main.bump_version", return_value="0.0.87"):
            cmd_align()

        self.assertEqual(ws.cell(shadok_row, col_map["en"]).value, before)
        # Non-shadok may be aligned (padding before colon)
        other_row = len(self.mapping) + 2
        self.assertNotEqual(ws.cell(other_row, col_map["en"]).value, "Other:x")

    @patch("src.main.open_or_create_workbook")
    def test_cmd_validate_shadok_integrity_counts_empty(
        self,
        mock_open_wb: MagicMock,
    ) -> None:
        wb = _make_workbook(self.mapping, self.lang_codes)  # all lang cells empty
        # Add a non-shadok complete row so structural validation has something
        ws = wb["Translations"]
        row = len(self.mapping) + 2
        ws.cell(row, 1, "Hello")
        for col, _lc in enumerate(self.lang_codes, 2):
            ws.cell(row, col, "Hello")
        mock_open_wb.return_value = wb

        import io

        buf = io.StringIO()
        with patch("src.main.load_languages", return_value=self.langs), \
             patch("src.main.get_shadok_target_langs", return_value=["en", "frca", "tr"]), \
             patch("src.main.BLOCK_JSON", ROOT / "data" / "missing_blocks_for_tests.json"), \
             patch("sys.stdout", buf):
            cmd_validate()

        out = buf.getvalue()
        self.assertIn("Shadok integrity", out)
        # One empty-block report per target lang
        self.assertGreaterEqual(out.count("empty block"), 3)
        col_map = _col_map(ws)
        resolved = resolve_shadok_mapping_rows(ws, col_map, self.mapping)
        self.assertEqual(len(resolved), 33)

    def test_shadok_prompt_strictness_escalates_per_attempt(self) -> None:
        p0 = build_shadok_system_prompt(0, 35, 39)
        p1 = build_shadok_system_prompt(
            1, 35, 39, previous_error="Expected at most 35 lines, got 36", previous_text="long"
        )
        p2 = build_shadok_system_prompt(
            2, 35, 39, previous_error="Line 1 visual_length 40 > 39", previous_text="bad"
        )
        self.assertIn("HARD SCREEN BUDGET", p0)
        self.assertIn("at most 35 lines", p0)
        self.assertIn("<= 39 characters", p0)
        self.assertNotIn("RETRY STRICTNESS LEVEL 1", p0)
        self.assertIn("RETRY STRICTNESS LEVEL 1", p1)
        self.assertIn("got 36", p1)
        self.assertIn("RETRY STRICTNESS LEVEL 2 (FINAL)", p2)
        self.assertIn("visual_length 40", p2)
        self.assertGreater(len(p2), len(p1))
        self.assertGreater(len(p1), len(p0))

    @patch("src.main.save_workbook")
    @patch("src.main.init_session_shadok")
    @patch("src.main.translate_shadok_block")
    @patch("src.main.open_or_create_workbook")
    def test_cmd_shadok_retries_then_succeeds(
        self,
        mock_open_wb: MagicMock,
        mock_translate: MagicMock,
        mock_init: MagicMock,
        mock_save: MagicMock,
    ) -> None:
        wb = _make_workbook(
            self.mapping,
            self.lang_codes,
            seed_lang_values={"en": "OLD"},
        )
        mock_open_wb.return_value = wb
        calls: list[int] = []

        def fake_block(
            full_text,
            target_langs,
            max_line_length,
            expected_lines=None,
            attempt=0,
            previous_error=None,
            previous_text=None,
        ):
            calls.append(attempt)
            if attempt == 0:
                self.assertIsNone(previous_error)
                return {"en": _valid_block("X", 36)}  # height overflow
            if attempt == 1:
                self.assertIn("at most 35", previous_error or "")
                self.assertIsNotNone(previous_text)
                return {"en": _valid_block("Y", 40)}  # still overflow
            self.assertEqual(attempt, 2)
            return {"en": _valid_block("Z", 28)}  # fewer is OK

        mock_translate.side_effect = fake_block

        with patch("src.main.get_shadok_target_langs", return_value=["en"]), \
             patch("sys.argv", ["main.py", "shadok"]):
            cmd_shadok()

        self.assertEqual(calls, [0, 1, 2])
        ws = wb["Translations"]
        col_map = _col_map(ws)
        self.assertEqual(ws.cell(2, col_map["en"]).value, "Z00 ok line")
        # Tail slots padded with a single space
        self.assertEqual(ws.cell(2 + 28, col_map["en"]).value, " ")
        mock_save.assert_called()

    def test_translate_shadok_uses_gemini_37(self) -> None:
        captured = {}

        def fake_retry(url, safe_data, payload, row_id=None, is_shadok=False, is_refine=False):
            captured["payload"] = payload
            return {"en": _valid_block("E", 3)}

        with patch.object(ai_client, "PROVIDER", "WEB2API"), \
             patch.object(ai_client, "_make_request_with_retry", side_effect=fake_retry):
            ai_client.translate_shadok_block("a\nb\nc", ["en"], 39, expected_lines=35)

        self.assertEqual(captured["payload"]["model"], ai_client.MODEL_WEB2API_SHADOK)
        self.assertEqual(ai_client.MODEL_WEB2API_SHADOK, "gemini-3.7-flash")

    def test_parse_json_safe_unescapes_newlines_in_fallback(self) -> None:
        # Unescaped interior " breaks json.loads; fallback must still decode \\n
        raw = '{"frca": "line one\\nline two"oops\\nline three"}'
        parsed = _parse_json_safe(raw)
        self.assertIn("frca", parsed)
        lines = parsed["frca"].split("\n")
        self.assertEqual(len(lines), 3)
        self.assertIn("oops", lines[1])


if __name__ == "__main__":
    unittest.main()
