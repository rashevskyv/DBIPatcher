"""Regression tests for PR #22 and PR #23 temperature aliases, cmd_sync deduplication, and workbook/CSV consistency."""

from __future__ import annotations

import ast
import csv
import json
import sys
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.core.text_utils import tokenize  # noqa: E402
from src.main import cmd_sync


class TemperatureAliasesAndSyncTests(unittest.TestCase):
    """Verify temperature aliases durability, single cmd_sync definition, and CSV consistency."""

    def setUp(self) -> None:
        self.dict_path = ROOT / "data" / "dictionary.xlsx"
        self.lang_path = ROOT / "data" / "languages.json"
        self.main_py_path = ROOT / "src" / "main.py"
        self.translations_dir = ROOT / "translations"

        with open(self.lang_path, "r", encoding="utf-8") as f:
            self.languages = json.load(f)

        self.wb = openpyxl.load_workbook(self.dict_path)
        self.ws = self.wb["Translations"]

        self.headers = [
            self.ws.cell(1, c).value for c in range(1, self.ws.max_column + 1)
        ]
        self.col_map = {h: i + 1 for i, h in enumerate(self.headers) if h}

    def tearDown(self) -> None:
        self.wb.close()

    def test_single_cmd_sync_definition(self) -> None:
        """Ensure exactly one cmd_sync function definition exists in src/main.py."""
        with open(self.main_py_path, "r", encoding="utf-8") as f:
            tree = ast.parse(f.read(), filename=str(self.main_py_path))

        sync_defs = [
            node.name
            for node in ast.walk(tree)
            if isinstance(node, ast.FunctionDef) and node.name == "cmd_sync"
        ]
        self.assertEqual(
            len(sync_defs),
            1,
            f"Expected exactly 1 definition of cmd_sync, found {len(sync_defs)}",
        )

    def test_workbook_structure_and_version(self) -> None:
        """Verify workbook contains 1,288 unique Original keys, 'tr' column, and version 0.0.90."""
        self.assertIn("tr", self.headers, "Translations sheet missing 'tr' column")

        meta_ws = self.wb["Metadata"]
        version = str(meta_ws["B1"].value or "")
        self.assertEqual(version, "0.0.90", f"Expected version 0.0.90, got {version}")

        original_col = self.col_map["Original"]
        originals = []
        for r in range(2, self.ws.max_row + 1):
            val = self.ws.cell(r, original_col).value
            if val is not None and str(val).strip():
                originals.append(str(val))

        self.assertEqual(
            len(originals),
            1288,
            f"Expected 1,288 original keys, found {len(originals)}",
        )
        self.assertEqual(
            len(set(originals)),
            1288,
            f"Expected 1,288 unique original keys, found {len(set(originals))}",
        )

    def test_all_45_aliases_and_canonical_values(self) -> None:
        """Verify all 45 PR #22 & PR #23 aliases exist once and match canonical translations."""
        can1_key = "Температура        : {}°C"
        can2_key = "Температура батареи             : {}°C"
        can3_key = "Средняя температура: {}°C"

        # 24 PR #22 literal '$°$' aliases
        pr22_aliases = [
            "Температура    : {}$°$C",
            "Температура     : {}$°$C",
            "Температура      : {}$°$C",
            "Температура       : {}$°$C",
            "Температура        : {}$°$C",
            "Температура         : {}$°$C",
            "Температура          : {}$°$C",
            "Температура           : {}$°$C",
            "Температура            : {}$°$C",
            "Температура             : {}$°$C",
            "Температура              : {}$°$C",
            "Срепняя температура: {}$°$C",
            "Средняя температура: {}$°$C",
            "Температура батареи        : {}$°$C",
            "Температура батареи         : {}$°$C",
            "Температура батареи          : {}$°$C",
            "Температура батареи           : {}$°$C",
            "Температура батареи            : {}$°$C",
            "Температура батареи             : {}$°$C",
            "Температура батареи              : {}$°$C",
            "Температура батареи               : {}$°$C",
            "Температура батареи                : {}$°$C",
            "Температура батареи                 : {}$°$C",
            "Температура батареи                  : {}$°$C",
        ]

        # 21 PR #23 clean '°' aliases
        pr23_aliases = [
            "Температура    : {}°C",
            "Температура     : {}°C",
            "Температура      : {}°C",
            "Температура       : {}°C",
            "Температура         : {}°C",
            "Температура          : {}°C",
            "Температура           : {}°C",
            "Температура            : {}°C",
            "Температура             : {}°C",
            "Температура              : {}°C",
            "Температура батареи        : {}°C",
            "Температура батареи         : {}°C",
            "Температура батареи          : {}°C",
            "Температура батареи           : {}°C",
            "Температура батареи            : {}°C",
            "Температура батареи              : {}°C",
            "Температура батареи               : {}°C",
            "Температура батареи                : {}°C",
            "Температура батареи                 : {}°C",
            "Температура батареи                  : {}°C",
            "Срепняя температура: {}°C",
        ]

        all_aliases = pr22_aliases + pr23_aliases
        self.assertEqual(len(all_aliases), 45)
        self.assertEqual(len(set(all_aliases)), 45)

        # Build row map
        original_col = self.col_map["Original"]
        key_to_row = {}
        for r in range(2, self.ws.max_row + 1):
            val = self.ws.cell(r, original_col).value
            if val is not None:
                key_to_row[str(val)] = r

        self.assertIn(can1_key, key_to_row)
        self.assertIn(can2_key, key_to_row)
        self.assertIn(can3_key, key_to_row)

        can1_row = key_to_row[can1_key]
        can2_row = key_to_row[can2_key]
        can3_row = key_to_row[can3_key]

        lang_cols = [lc for lc in self.languages if lc in self.col_map]
        self.assertIn("tr", lang_cols)

        for alias in all_aliases:
            self.assertIn(alias, key_to_row, f"Missing alias in workbook: {repr(alias)}")
            alias_row = key_to_row[alias]

            if alias.startswith("Температура батареи"):
                expected_src_row = can2_row
            elif alias.startswith("Средняя температура") or alias.startswith("Срепняя температура"):
                expected_src_row = can3_row
            elif alias.startswith("Температура"):
                expected_src_row = can1_row
            else:
                self.fail(f"Unrecognized alias prefix: {repr(alias)}")

            for lc in lang_cols:
                col_idx = self.col_map[lc]
                actual_val = self.ws.cell(alias_row, col_idx).value
                expected_val = self.ws.cell(expected_src_row, col_idx).value
                if lc == "tr":
                    # tr canonical row received UI alignment spaces in v0.0.89 while unaligned aliases retained tokens
                    self.assertEqual(
                        str(actual_val).split(),
                        str(expected_val).split(),
                        f"Mismatch in alias {repr(alias)} for language '{lc}': got {repr(actual_val)}, expected {repr(expected_val)}",
                    )
                else:
                    self.assertEqual(
                        actual_val,
                        expected_val,
                        f"Mismatch in alias {repr(alias)} for language '{lc}': got {repr(actual_val)}, expected {repr(expected_val)}",
                    )

    def test_exported_csv_keys_match_workbook(self) -> None:
        """Verify every configured language CSV matches the workbook key set exactly."""
        original_col = self.col_map["Original"]
        wb_tokenized_keys = set()
        for r in range(2, self.ws.max_row + 1):
            val = self.ws.cell(r, original_col).value
            if val is not None and str(val).strip():
                wb_tokenized_keys.add(str(val))

        self.assertEqual(len(wb_tokenized_keys), 1288)

        for lc in self.languages:
            if lc == "ru":
                continue
            csv_path = self.translations_dir / f"{lc}.csv"
            self.assertTrue(csv_path.exists(), f"Missing CSV file: {csv_path}")

            csv_keys = set()
            with open(csv_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                header = next(reader)
                self.assertEqual(header, ["original", "translation"], f"Bad header in {csv_path}")
                for row in reader:
                    if row:
                        csv_keys.add(tokenize(row[0]))

            self.assertEqual(
                csv_keys,
                wb_tokenized_keys,
                f"Mismatch in source keys for {lc}.csv (diff size: {len(csv_keys ^ wb_tokenized_keys)})",
            )

    def test_cmd_sync_deduplication_and_cell_merging(self) -> None:
        """Verify cmd_sync applied to A, B, A, B, C retains exactly A, B, C and merges complementary cells."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Translations"
        meta = wb.create_sheet("Metadata")
        meta["A1"] = "version"
        meta["B1"] = "0.0.87"
        meta["A2"] = "updated"
        meta["B2"] = ""

        # Columns: Original, ru, ua, en, de
        headers = ["Original", "ru", "ua", "en", "de"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(1, col_idx, h)

        # Initial rows:
        # Row 2: A, ru='A_ru', ua='A_ua', en='A_en_first', de=None
        # Row 3: B, ru='B_ru', ua=None, en=None, de='B_de_first'
        # Row 4: A, ru=None, ua=None, en='A_en_second', de='A_de_second'
        # Row 5: B, ru=None, ua='B_ua_second', en='B_en_second', de=None
        # Row 6: C, ru='C_ru', ua='C_ua', en='C_en', de='C_de'
        rows_data = [
            ["A", "A_ru", "A_ua", "A_en_first", None],
            ["B", "B_ru", None, None, "B_de_first"],
            ["A", None, None, "A_en_second", "A_de_second"],
            ["B", None, "B_ua_second", "B_en_second", None],
            ["C", "C_ru", "C_ua", "C_en", "C_de"],
        ]
        for r_idx, row in enumerate(rows_data, 2):
            for c_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(r_idx, c_idx, val)

        with patch("src.main.open_or_create_workbook", return_value=wb), \
             patch("src.main.save_workbook") as mock_save, \
             patch("src.main.load_languages", return_value={"ru": "Russian", "ua": "Ukrainian", "en": "English", "de": "German"}), \
             patch("src.main.DATA_DIR", ROOT / "non_existent_data_dir"):
            cmd_sync()

        mock_save.assert_called_once_with(wb)

        # Header (row 1) + 3 retained data rows (A, B, C) -> max_row == 4
        self.assertEqual(ws.max_row, 4)

        # Verify Row 2 (A)
        self.assertEqual(ws.cell(2, 1).value, "A")
        self.assertEqual(ws.cell(2, 2).value, "A_ru")
        self.assertEqual(ws.cell(2, 3).value, "A_ua")
        self.assertEqual(ws.cell(2, 4).value, "A_en_first")  # First row value retained when both non-empty and differ
        self.assertEqual(ws.cell(2, 5).value, "A_de_second") # Empty cell merged from duplicate row

        # Verify Row 3 (B)
        self.assertEqual(ws.cell(3, 1).value, "B")
        self.assertEqual(ws.cell(3, 2).value, "B_ru")
        self.assertEqual(ws.cell(3, 3).value, "B_ua_second") # Merged from duplicate row
        self.assertEqual(ws.cell(3, 4).value, "B_en_second") # Merged from duplicate row
        self.assertEqual(ws.cell(3, 5).value, "B_de_first")  # Retained from first row

        # Verify Row 4 (C) - Unrelated unique row was not deleted
        self.assertEqual(ws.cell(4, 1).value, "C")
        self.assertEqual(ws.cell(4, 2).value, "C_ru")
        self.assertEqual(ws.cell(4, 3).value, "C_ua")
        self.assertEqual(ws.cell(4, 4).value, "C_en")
        self.assertEqual(ws.cell(4, 5).value, "C_de")


if __name__ == "__main__":
    unittest.main()
