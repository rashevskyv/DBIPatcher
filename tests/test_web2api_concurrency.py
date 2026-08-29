"""Focused unit and concurrency tests for Web2API provider and cmd_translate worker pool."""

from __future__ import annotations

import json
import os
import sys
import tempfile
import threading
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.core import ai_client
from src.core.ai_client import (
    MODEL_WEB2API,
    WEB2API_MODELS_URL,
    _log_interaction,
    _make_request_with_retry,
    check_web2api_preflight,
    refine,
)
from src.main import (
    RowTranslationResult,
    _process_row_translation,
    cmd_translate,
    get_translate_workers,
)


class Web2APIConcurrencyTests(unittest.TestCase):
    """Test suite covering Web2API bounded concurrency, worker limits, and stateless retry/refine."""

    def setUp(self) -> None:
        self.orig_provider = ai_client.PROVIDER
        self.orig_env_workers = os.environ.get("DBI_TRANSLATE_WORKERS")

    def tearDown(self) -> None:
        ai_client.PROVIDER = self.orig_provider
        if self.orig_env_workers is not None:
            os.environ["DBI_TRANSLATE_WORKERS"] = self.orig_env_workers
        else:
            os.environ.pop("DBI_TRANSLATE_WORKERS", None)

    def test_worker_limits_validation(self) -> None:
        """Verify DBI_TRANSLATE_WORKERS default and range 1..8 validation."""
        # Default when unset
        os.environ.pop("DBI_TRANSLATE_WORKERS", None)
        self.assertEqual(get_translate_workers(), 4)

        # Valid range
        for valid in ["1", "2", "4", "8"]:
            os.environ["DBI_TRANSLATE_WORKERS"] = valid
            self.assertEqual(get_translate_workers(), int(valid))

        # Invalid values must raise ValueError
        for invalid in ["0", "9", "-1", "100", "abc", "4.5", ""]:
            os.environ["DBI_TRANSLATE_WORKERS"] = invalid
            with self.assertRaises(ValueError, msg=f"Expected ValueError for {invalid!r}"):
                get_translate_workers()

    @patch("src.core.ai_client.requests.get")
    def test_web2api_preflight_success_and_failure(self, mock_get: MagicMock) -> None:
        """Verify check_web2api_preflight parses /v1/models and verifies MODEL_WEB2API."""
        # 1. Success: model is present
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "data": [
                {"id": "gemini-3.7-flash"},
                {"id": MODEL_WEB2API},
                {"id": "gemini-3.5-flash"},
            ]
        }
        mock_get.return_value = mock_resp

        # Should not raise
        check_web2api_preflight()
        mock_get.assert_called_once_with(WEB2API_MODELS_URL, timeout=10)

        # 2. Failure: model is missing from data[].id
        mock_resp.json.return_value = {"data": [{"id": "other-model"}]}
        with self.assertRaises(RuntimeError) as ctx:
            check_web2api_preflight()
        self.assertIn(MODEL_WEB2API, str(ctx.exception))

        # 3. Failure: non-200 status code
        mock_resp.status_code = 503
        with self.assertRaises(RuntimeError) as ctx:
            check_web2api_preflight()
        self.assertIn("503", str(ctx.exception))

    @patch("src.core.ai_client.init_session")
    @patch("src.core.ai_client.requests.post")
    def test_web2api_retry_does_not_call_init_session_or_clear_log(
        self,
        mock_post: MagicMock,
        mock_init_session: MagicMock,
    ) -> None:
        """Verify Web2API failure retries at most once and never calls init_session on failure."""
        ai_client.PROVIDER = "WEB2API"

        mock_resp = MagicMock()
        mock_resp.status_code = 500
        mock_resp.text = "Internal Server Error"
        mock_post.return_value = mock_resp

        payload = {"model": MODEL_WEB2API, "messages": []}
        safe_data = json.dumps(payload).encode("utf-8")

        with patch("time.sleep"):
            with self.assertRaises(RuntimeError):
                _make_request_with_retry(
                    "http://localhost:8081/v1/chat/completions",
                    safe_data,
                    payload,
                    row_id=1,
                )

        # At most 1 client-side retry -> 2 total calls
        self.assertEqual(mock_post.call_count, 2)
        # Must NEVER re-initialize session / clear logs for WEB2API during worker execution
        mock_init_session.assert_not_called()

    @patch("src.core.ai_client.init_session")
    @patch("src.core.ai_client.requests.post")
    def test_omniroad_retry_performs_session_recovery(
        self,
        mock_post: MagicMock,
        mock_init_session: MagicMock,
    ) -> None:
        """Verify OmniRoad provider retains its existing session recovery attempt on final retry failure."""
        ai_client.PROVIDER = "OMNIROAD"

        # 3 failure responses (attempt 0, 1, 2) followed by 1 successful recovery response
        fail_resp = MagicMock()
        fail_resp.status_code = 500
        fail_resp.text = "Internal Server Error"

        success_resp = MagicMock()
        success_resp.status_code = 200
        success_resp.text = json.dumps({
            "choices": [{"message": {"content": json.dumps({"en": "Recovered"})}}]
        })

        mock_post.side_effect = [fail_resp, fail_resp, fail_resp, success_resp]

        payload = {"model": ai_client.MODEL_OMNI, "messages": []}
        safe_data = json.dumps(payload).encode("utf-8")

        with patch("time.sleep"):
            res = _make_request_with_retry(
                ai_client.OMNIROAD_URL,
                safe_data,
                payload,
                row_id=1,
            )

        self.assertEqual(res, {"en": "Recovered"})
        # 3 regular attempts + 1 recovery attempt = 4 calls total
        self.assertEqual(mock_post.call_count, 4)
        mock_init_session.assert_called_once()

    @patch("src.core.ai_client.requests.post")
    def test_stateless_refine_payload_structure(self, mock_post: MagicMock) -> None:
        """Verify refine builds a stateless payload with system prompt, source text, candidate values, and errors."""
        ai_client.PROVIDER = "WEB2API"

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.text = json.dumps({"choices": [{"message": {"content": json.dumps({"en": "Start"})}}]})
        mock_post.return_value = mock_resp

        result = refine(
            correction="Fix colon alignment",
            target_langs=["en"],
            row_id=10,
            original="Запуск :",
            current_translations={"en": "Start"},
            validation_errors=[("en", "Missing colon")],
        )

        self.assertEqual(result, {"en": "Start"})
        self.assertEqual(mock_post.call_count, 1)

        call_kwargs = mock_post.call_args[1]
        sent_data = json.loads(call_kwargs["data"].decode("utf-8"))

        messages = sent_data["messages"]
        self.assertEqual(len(messages), 2)
        self.assertEqual(messages[0]["role"], "system")
        self.assertEqual(messages[1]["role"], "user")

        user_content = messages[1]["content"]
        self.assertIn("Запуск :", user_content)
        self.assertIn("Start", user_content)
        self.assertIn("Missing colon", user_content)
        self.assertIn("Fix colon alignment", user_content)
        self.assertIn("en", user_content)

    @patch("src.main.refine")
    @patch("src.main.translate_batch")
    def test_pure_row_worker_does_not_touch_workbook(
        self,
        mock_translate: MagicMock,
        mock_refine: MagicMock,
    ) -> None:
        """Verify _process_row_translation is pure and returns RowTranslationResult without mutating workbook."""
        mock_translate.return_value = {
            "en": "File [[LF]] saved",
            "de": "Datei [[LF]] gespeichert",
        }

        res = _process_row_translation(
            row_id=42,
            original="Файл [[LF]] сохранен",
            missing=["en", "de"],
        )

        self.assertIsInstance(res, RowTranslationResult)
        self.assertEqual(res.row_id, 42)
        self.assertEqual(res.original, "Файл [[LF]] сохранен")
        self.assertEqual(res.accepted_translations["en"], "File [[LF]] saved")
        self.assertEqual(res.accepted_translations["de"], "Datei [[LF]] gespeichert")
        self.assertEqual(res.failed_langs, [])
        self.assertFalse(res.is_skipped_cyrillic)

    def test_pure_row_worker_skips_cyrillic_under_two(self) -> None:
        """Verify rows with less than two Cyrillic characters are copied deterministically."""
        res = _process_row_translation(
            row_id=5,
            original="TitleId : {} BID: {}",
            missing=["en", "de"],
        )
        self.assertTrue(res.is_skipped_cyrillic)
        self.assertEqual(res.accepted_translations["en"], "TitleId : {} BID: {}")
        self.assertEqual(res.accepted_translations["de"], "TitleId : {} BID: {}")

    @patch("src.main.save_workbook")
    @patch("src.main.init_session")
    @patch("src.main.open_or_create_workbook")
    @patch("src.main._process_row_translation")
    def test_concurrent_execution_in_flight_for_web2api(
        self,
        mock_process_row: MagicMock,
        mock_open_wb: MagicMock,
        mock_init_session: MagicMock,
        mock_save_wb: MagicMock,
    ) -> None:
        """Verify that with WEB2API and workers greater than 1, multiple rows execute concurrently."""
        ai_client.PROVIDER = "WEB2API"
        os.environ["DBI_TRANSLATE_WORKERS"] = "2"

        # Set up mock workbook with 2 rows needing translation
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_wb.__getitem__.return_value = mock_ws
        mock_open_wb.return_value = mock_wb

        # Header: Original, en, de
        mock_ws.max_row = 3
        mock_ws.max_column = 3

        def get_cell(row, col, value=None):
            cell = MagicMock()
            if row == 1:
                cell.value = ["Original", "en", "de"][col - 1]
            elif row == 2:
                cell.value = "Строка 1" if col == 1 else None
            elif row == 3:
                cell.value = "Строка 2" if col == 1 else None
            return cell

        mock_ws.cell.side_effect = get_cell

        # Synchronize 2 concurrent in-flight worker calls using Barrier
        barrier = threading.Barrier(2, timeout=5)
        in_flight_max = 0
        current_in_flight = 0
        flight_lock = threading.Lock()

        def concurrent_worker(row_id, original, missing):
            nonlocal in_flight_max, current_in_flight
            with flight_lock:
                current_in_flight += 1
                if current_in_flight > in_flight_max:
                    in_flight_max = current_in_flight
            # Wait until both workers are in flight at the same time
            try:
                barrier.wait()
            finally:
                with flight_lock:
                    current_in_flight -= 1
            return RowTranslationResult(
                row_id=row_id,
                original=original,
                missing=missing,
                accepted_translations={lc: f"{original}_{lc}" for lc in missing},
            )

        mock_process_row.side_effect = concurrent_worker

        with patch("src.main.load_languages", return_value={"en": "English", "de": "German"}):
            with patch("src.main.bump_version", return_value="0.0.87"):
                with patch("sys.argv", ["main.py", "translate"]):
                    cmd_translate()

        # Assert that at least 2 workers were in flight simultaneously
        self.assertGreaterEqual(in_flight_max, 2)
        # Assert main thread saved workbook after each completed row (2 saves) + 1 final version bump save
        self.assertGreaterEqual(mock_save_wb.call_count, 3)

    @patch("src.main.save_workbook")
    @patch("src.main.init_session")
    @patch("src.main.open_or_create_workbook")
    @patch("src.main._process_row_translation")
    def test_serial_execution_for_legacy_providers(
        self,
        mock_process_row: MagicMock,
        mock_open_wb: MagicMock,
        mock_init_session: MagicMock,
        mock_save_wb: MagicMock,
    ) -> None:
        """Verify GEMINI_PROXY and OMNIROAD remain strictly serial even if workers=4."""
        for provider in ["GEMINI_PROXY", "OMNIROAD"]:
            with self.subTest(provider=provider):
                ai_client.PROVIDER = provider
                os.environ["DBI_TRANSLATE_WORKERS"] = "4"

                mock_wb = MagicMock()
                mock_ws = MagicMock()
                mock_wb.__getitem__.return_value = mock_ws
                mock_open_wb.return_value = mock_wb

                mock_ws.max_row = 3
                mock_ws.max_column = 2

                def get_cell(row, col, value=None):
                    cell = MagicMock()
                    if row == 1:
                        cell.value = "Original" if col == 1 else "en"
                    elif row in (2, 3):
                        cell.value = f"Строка {row}" if col == 1 else None
                    return cell

                mock_ws.cell.side_effect = get_cell

                active_threads = set()
                thread_lock = threading.Lock()

                def serial_worker(row_id, original, missing):
                    with thread_lock:
                        active_threads.add(threading.current_thread().name)
                    return RowTranslationResult(
                        row_id=row_id,
                        original=original,
                        missing=missing,
                        accepted_translations={"en": f"{original}_en"},
                    )

                mock_process_row.side_effect = serial_worker

                with patch("src.main.load_languages", return_value={"en": "English"}):
                    with patch("src.main.bump_version", return_value="0.0.87"):
                        with patch("sys.argv", ["main.py", "translate"]):
                            cmd_translate()

                # In serial execution, all work executes on the MainThread
                self.assertEqual(active_threads, {"MainThread"})

    def test_log_append_thread_safety(self) -> None:
        """Verify concurrent _log_interaction calls safely append to a temporary log file without data loss."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_log_path = Path(tmp_dir) / "test_ai_proxy.log"

            with patch("src.core.ai_client.LOG_FILE", temp_log_path):
                threads = []
                entry_count = 10
                for i in range(entry_count):
                    t = threading.Thread(
                        target=_log_interaction,
                        args=({"worker_id": i, "payload": f"payload_{i}"}, f"response_{i}", i),
                    )
                    threads.append(t)
                    t.start()

                for t in threads:
                    t.join()

                self.assertTrue(temp_log_path.exists(), "Expected temporary log file to be created")
                content = temp_log_path.read_text(encoding="utf-8")

                for i in range(entry_count):
                    self.assertEqual(
                        content.count(f"ROW: {i}"),
                        1,
                        f"Expected exactly one 'ROW: {i}' marker in log",
                    )
                    self.assertEqual(
                        content.count(f"response_{i}"),
                        1,
                        f"Expected exactly one 'response_{i}' entry in log",
                    )
                    self.assertIn(f'"worker_id": {i}', content)

    @patch("src.main.save_workbook")
    @patch("src.main.init_session")
    @patch("src.main.open_or_create_workbook")
    @patch("src.main.translate_batch")
    def test_cmd_translate_retains_invalid_value_on_failure_and_checkpoints(
        self,
        mock_translate_batch: MagicMock,
        mock_open_wb: MagicMock,
        mock_init_session: MagicMock,
        mock_save_wb: MagicMock,
    ) -> None:
        """Verify pre-existing invalid translation is kept on failure and preserved across all checkpoints."""
        ai_client.PROVIDER = "WEB2API"
        os.environ["DBI_TRANSLATE_WORKERS"] = "2"

        # Create in-memory workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Translations"
        meta = wb.create_sheet("Metadata")
        meta["A1"] = "version"
        meta["B1"] = "0.0.87"
        meta["A2"] = "updated"
        meta["B2"] = ""

        # Columns: Original, en, de
        ws.cell(1, 1, "Original")
        ws.cell(1, 2, "en")
        ws.cell(1, 3, "de")

        # Row 2: Failed translation row (has pre-existing invalid translation in 'en')
        ws.cell(2, 1, "Строка {} с ошибкой:")
        ws.cell(2, 2, "Old Invalid without colon or placeholder")
        ws.cell(2, 3, "Gültig {} Deutsch:")

        # Row 3: Successful translation row (has pre-existing invalid translation in 'en')
        ws.cell(3, 1, "Успешная {} строка:")
        ws.cell(3, 2, "Another Bad without colon or placeholder")
        ws.cell(3, 3, "Auch {} Deutsch:")

        mock_open_wb.return_value = wb

        expected_row3_en = "Valid Успешная {} строка:"

        # Mock translate_batch:
        # Row 2 raises Exception to simulate complete translation failure
        # Row 3 returns valid translation
        def mock_batch(original, missing, row_id=None):
            if "ошибкой" in original:
                raise RuntimeError("AI service unavailable for Row 2")
            return {lc: f"Valid {original}" for lc in missing}

        mock_translate_batch.side_effect = mock_batch

        # Track cell values at each save_workbook checkpoint
        checkpoints = []
        def record_checkpoint(saved_wb):
            ws_check = saved_wb["Translations"]
            checkpoints.append({
                "row2_en": ws_check.cell(2, 2).value,
                "row3_en": ws_check.cell(3, 2).value,
            })
        mock_save_wb.side_effect = record_checkpoint

        with patch("src.main.load_languages", return_value={"en": "English", "de": "German"}), \
             patch("src.main.bump_version", return_value="0.0.87"), \
             patch("sys.argv", ["main.py", "translate"]):
            cmd_translate()

        # Exactly two checkpoints: one after Row 3 successfully applies and one final save
        self.assertEqual(len(checkpoints), 2)
        self.assertEqual(checkpoints[0]["row3_en"], expected_row3_en)
        self.assertEqual(checkpoints[1]["row3_en"], expected_row3_en)

        # In EVERY checkpoint, row 2 must retain its old invalid translation "Old Invalid without colon or placeholder"
        for cp in checkpoints:
            self.assertEqual(
                cp["row2_en"],
                "Old Invalid without colon or placeholder",
                f"Row 2 lost its pre-existing value in checkpoint: {cp}",
            )

        # Final state in workbook:
        # Row 2 failed -> retained pre-existing translation
        self.assertEqual(ws.cell(2, 2).value, "Old Invalid without colon or placeholder")
        # Row 3 succeeded -> replaced with new valid translation
        self.assertEqual(ws.cell(3, 2).value, expected_row3_en)


if __name__ == "__main__":
    unittest.main()
